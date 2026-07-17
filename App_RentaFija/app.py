# -*- coding: utf-8 -*-
"""
Curvas de ONs y Bonos — versión PRO v2 (FIX de datasets + Carry Trade)
-----------------------------------------------------------------------
CAMBIOS RESPECTO A LA VERSIÓN ANTERIOR (por qué el breakeven no andaba bien):

  1) DATASET_RF_SOBERANA (antes 42760, "Historical Price - Spot") NO tenía columnas
     de TIR ni Modified Duration -- solo price/volume. Se elimina: los soberanos en
     pesos (CER/Fija/DL/Duales) ahora se toman del MISMO dataset 41886 que ya usan
     las ONs y los bonos HD, filtrando por "market segment" == "Sovereign". Ese
     dataset SÍ trae irr/modified duration para todo el universo.

  2) DATASET_REM (antes 5621) apuntaba a "Monthly GDP (IMAEP) - Banco Central del
     Paraguay": el PBI mensual de Paraguay, sin ninguna relación con expectativas de
     inflación de Argentina. Se reemplaza por el dataset 44033
     ("Inflation Expectations (REM)", BCRA), que trae Variable/Referencia/Período/
     Mediana y permite filtrar directamente "Próx. 12 meses" para la inflación
     interanual esperada.

  3) DATASET_FUTUROS (antes 5331) apuntaba al EMAE (INDEC) -- actividad económica,
     no futuros de dólar. Se reemplaza por el dataset 5361 ("Dollar Futures -
     Estimated Implied Curve", repositorio Matba Rofex), que trae Spot y la curva
     interpolada a 30/60/.../330 días.

  4) La clasificación CER/Fija/DL por REGEX sobre el ticker se reemplaza por el
     campo "coupon structure" que ya viene provisto y clasificado en el dataset
     (mucho más confiable). Se deja la posibilidad de reclasificar manualmente
     por si el usuario quiere mover algún caso borde (ej. Duales).

  5) Se agrega una sección explícita de CARRY TRADE: para cada bono CER/DL se
     calcula el carry (en bps) entre el breakeven que pricea el mercado y la
     expectativa de REM (para CER) o de la curva de futuros ROFEX (para DL),
     interpolada a la duration de cada bono.

Requisitos:
  pip install streamlit pandas numpy plotly scipy alphacast openpyxl
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from alphacast import Alphacast
from scipy.optimize import curve_fit
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, LineChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule

# =====================================================================
# Configuración general y tema visual
# =====================================================================

DEFAULT_DATASET_ID = 41886        # ONs / Bonos / Soberanos (curvas) -- TODO el universo
DATASET_REM = 44033                # REM - BCRA - Inflation Expectations (CORREGIDO)
DATASET_FUTUROS = 5361             # Dollar Futures - Estimated Implied Curve (CORREGIDO)
# DATASET_RF_SOBERANA eliminado: los soberanos en pesos salen del dataset principal.

COLORS = {
    "primary": "#1f2a44",     # azul institucional
    "accent": "#c8963e",      # dorado
    "gd": "#2563eb",          # ley NY
    "al": "#dc2626",          # ley local
    "bopreal": "#059669",     # BCRA
    "cer": "#7c3aed",
    "fija": "#0891b2",
    "dl": "#ea580c",
    "dual": "#a16207",
    "cheap": "#16a34a",
    "rich": "#dc2626",
    "grid": "rgba(120,130,150,0.18)",
    "fit": "#64748b",
}

PLOTLY_LAYOUT = dict(
    template="plotly_white",
    font=dict(family="Segoe UI, Helvetica, Arial", size=13, color="#1e293b"),
    title_font=dict(size=17, color=COLORS["primary"]),
    margin=dict(l=60, r=30, t=70, b=55),
    hoverlabel=dict(bgcolor="#1e293b", font=dict(size=12, color="white"), bordercolor="#1e293b"),
    legend=dict(orientation="h", yanchor="bottom", y=1.01, xanchor="right", x=1,
                bgcolor="rgba(255,255,255,0.6)"),
    plot_bgcolor="white",
)


def style_axes(fig: go.Figure, xtitle: str, ytitle: str) -> go.Figure:
    fig.update_xaxes(title=xtitle, showgrid=True, gridcolor=COLORS["grid"],
                     zeroline=False, showline=True, linecolor="#94a3b8", ticks="outside")
    fig.update_yaxes(title=ytitle, showgrid=True, gridcolor=COLORS["grid"],
                     zeroline=False, showline=True, linecolor="#94a3b8", ticks="outside")
    return fig


def add_source_watermark(fig: go.Figure, fecha=None, fuente="Alphacast") -> go.Figure:
    txt = f"Fuente: {fuente}"
    if fecha is not None:
        txt += f" · Datos al {pd.Timestamp(fecha).strftime('%d/%m/%Y')}"
    fig.add_annotation(text=txt, xref="paper", yref="paper", x=0, y=-0.16,
                       showarrow=False, font=dict(size=10, color="#94a3b8"), align="left")
    return fig


# =====================================================================
# Universos por defecto
# =====================================================================

DEFAULT_ONS = {
    "YMCHO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "MGC9O": {"calificacion": "AA+(arg)", "min_nominal": 1},
    "MTCGO": {"calificacion": "N/D", "min_nominal": 1},
    "TLC1O": {"calificacion": "AA+(arg)", "min_nominal": 1000},
    "PNDCO": {"calificacion": "AAA(arg)", "min_nominal": 1000},
    "GNCXO": {"calificacion": "A+(arg)", "min_nominal": 1000},
    "BACAO": {"calificacion": "AAA(arg)", "min_nominal": 150000},
    "CAC5O": {"calificacion": "AA(arg)", "min_nominal": 1},
    "YCAMO": {"calificacion": "AAA(arg)", "min_nominal": 10000},
    "IRCFO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "YMCIO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "BYCHO": {"calificacion": "A1+(arg)", "min_nominal": 2000},
    "ARC1O": {"calificacion": "AA(arg)", "min_nominal": 1000},
    "BACGO": {"calificacion": "CCC+ (Fitch)", "min_nominal": 1000},
    "PNXCO": {"calificacion": "AAA(arg)", "min_nominal": 1000},
    "DNC7O": {"calificacion": "CCC+ (Fitch)", "min_nominal": 100},
    "RUCDO": {"calificacion": "CCC (Fitch)", "min_nominal": 1},
    "TLCMO": {"calificacion": "AA+(arg)", "min_nominal": 1000},
    "YMCXO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "TSC3O": {"calificacion": "B (Fitch)", "min_nominal": 10000},
    "MGCMO": {"calificacion": "AA+(arg)", "min_nominal": 1000},
    "YFCJO": {"calificacion": "AAA(arg)", "min_nominal": 1000},
    "PLC4O": {"calificacion": "BB (Fitch)", "min_nominal": 1000},
    "YMCJO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "TTCAO": {"calificacion": "AAA(arg)", "min_nominal": 1000},
    "VSCVO": {"calificacion": "BB- (Fitch)", "min_nominal": 1000},
    "TLCPO": {"calificacion": "B (Fitch)", "min_nominal": 100},
    "RC1CO": {"calificacion": "B+ (Fitch)", "min_nominal": 1000},
    "YM34O": {"calificacion": "AAA(arg)", "min_nominal": 1000},
    "IRCPO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "MGCOO": {"calificacion": "AAA(arg)", "min_nominal": 10000},
    "VSCTO": {"calificacion": "AAA(arg)", "min_nominal": 10000},
    "DNC5O": {"calificacion": "A(arg)", "min_nominal": 1},
    "LOC5O": {"calificacion": "AAA(arg)", "min_nominal": 50},
    "PN38O": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "RCCJO": {"calificacion": "AAA(arg)", "min_nominal": "N/D"},
    "VSCRO": {"calificacion": "AAA(arg)", "min_nominal": 1},
    "DNC3O":  {"calificacion": "A+(arg)",  "min_nominal": 1},        # Edenor 2026
    "PN35O":  {"calificacion": "AAA(arg)", "min_nominal": 1},        # PAE 2029
    "TTCDO":  {"calificacion": "AAA(arg)", "min_nominal": 10000},    # Tecpetrol 2030
    "PLC5O":  {"calificacion": "AAA(arg)", "min_nominal": 1000},     # PlusPetrol 2031
    "DNCAO":  {"calificacion": "A+(arg)",  "min_nominal": 100},      # Edenor 2033
    "GN49O":  {"calificacion": "A+(arg)",  "min_nominal": 1000},     # Genneia 2033
    "TLCTO":  {"calificacion": "AA+(arg)", "min_nominal": 1},        # Telecom 2036
    "PN43O":  {"calificacion": "AAA(arg)", "min_nominal": 1000},     # PAE 2037
    "TSC4O":  {"calificacion": "B- (Fitch)", "min_nominal": 1000},   # TGS 2035 (Global -> Fitch, mismo criterio que TSC3O)
    "VSCXO":  {"calificacion": "AAA(arg)", "min_nominal": 1000},     # Vista Oil & Gas 2038
    "MGCRO":  {"calificacion": "AAA(arg)", "min_nominal": 1000},     # Pampa Energía 2037
}

DEFAULT_LAMINA_MAX = 10000
DEFAULT_MD_MAX = 5.0

BONOS_GD = ["GD29", "GD30", "GD35", "GD38", "GD41", "GD46"]
BONOS_AL = ["AL29", "AL30", "AL35", "AE38", "AL41", "AL46"]
BOPREAL = ["BPOC7", "BPOC8", "BPOC9"]

# Clasificación de soberanos en pesos a partir de la columna "coupon structure"
# que ya viene provista por el dataset (reemplaza los regex frágiles sobre ticker).
COUPON_TO_CLASE = {
    "ARS inflation-linked rate": "CER",
    "ARS fixed rate": "Fija",
    "Dollar-linked rate": "DL",
    "Dual (CER Dollar-linked rate)": "Dual",
    "Dual (Fixed or TAMAR rate)": "Dual",
    "Dual (CER or TAMAR rate)": "Dual",
    "ARS floating rate": "Badlar/Pase",
}


# =====================================================================
# Descarga y normalización
# =====================================================================

def default_ons_df() -> pd.DataFrame:
    rows = [{"Ticker": t, "Calificacion": m.get("calificacion", "N/D"),
             "Lamina_Minima": m.get("min_nominal", np.nan)} for t, m in DEFAULT_ONS.items()]
    df = pd.DataFrame(rows)
    df["Lamina_Minima"] = pd.to_numeric(df["Lamina_Minima"], errors="coerce")
    return df


@st.cache_data(show_spinner=False, ttl=15 * 60)
def download_dataset(api_key: str, dataset_id: int) -> pd.DataFrame:
    alphacast = Alphacast(api_key)
    csv_bytes = alphacast.datasets.dataset(int(dataset_id)).download_data(format="csv")
    return pd.read_csv(io.StringIO(csv_bytes.decode("utf-8")))


def normalize_dataset(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza el dataset principal (41886). A diferencia de la versión anterior,
    CONSERVA 'market segment' y 'coupon structure': se necesitan para clasificar
    soberanos en pesos (CER/Fija/DL/Dual) sin depender de regex sobre el ticker."""
    df = df.copy()

    ren = {"symbol": "Ticker", "irr": "TIR", "modified duration": "MD",
           "convexity": "Convexidad", "parity": "Paridad", "residual value": "Valor Residual",
           "market segment": "Segmento", "coupon structure": "CouponStructure",
           "issue currency": "IssueCcy", "trading currency": "TradingCcy",
           "volume": "Volumen", "issuer": "Emisor"}
    df.rename(columns={k: v for k, v in ren.items() if k in df.columns}, inplace=True)

    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    if "TIR" in df.columns:
        df["TIR"] = pd.to_numeric(df["TIR"], errors="coerce") * 100
    if "MD" in df.columns:
        df["MD"] = pd.to_numeric(df["MD"], errors="coerce")
    if "Volumen" in df.columns:
        df["Volumen"] = pd.to_numeric(df["Volumen"], errors="coerce")

    wanted = ["Date", "Ticker", "Industry", "Emisor", "law", "Segmento", "CouponStructure",
              "IssueCcy", "TradingCcy", "TIR", "MD", "Convexidad", "Paridad", "Valor Residual", "Volumen"]
    df = df[[c for c in wanted if c in df.columns]].copy()
    for c in ["TIR", "MD", "Convexidad", "Paridad"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)
    return df


def latest_snapshot(df_norm: pd.DataFrame):
    if "Date" not in df_norm.columns:
        return df_norm.copy(), None
    d = df_norm["Date"].max()
    return df_norm[df_norm["Date"] == d].copy(), d


def build_ons_table(latest_df: pd.DataFrame, ons_df: pd.DataFrame) -> pd.DataFrame:
    if latest_df.empty:
        return latest_df.copy()
    df = latest_df.copy()
    df["Ticker"] = df["Ticker"].astype(str)
    o = ons_df.copy()
    o["Ticker"] = o["Ticker"].astype(str)
    o["Lamina_Minima"] = pd.to_numeric(o["Lamina_Minima"], errors="coerce")
    df = df[df["Ticker"].isin(o["Ticker"].dropna().unique())].copy()
    return df.merge(o[["Ticker", "Calificacion", "Lamina_Minima"]], on="Ticker", how="left")


# Bonos de las reestructuraciones 2005/2010 (Par/Discount/Cuasipar): son performing
# pero suelen ser ilíquidos y con estructura de cupón atípica que ensucia el fit de
# la curva CER moderna. Se excluyen por LIQUIDEZ/COMPARABILIDAD, no por default.
LEGACY_PREFIXES = ("PARP", "DICP", "CUAP", "DIP0", "PAP0", "PARY", "PAY0", "TVY0", "DICPD", "CUAPC")


def classify_soberanos_pesos(df_norm: pd.DataFrame, min_md: float = 0.05, max_md: float = 6.0,
                             drop_legacy: bool = True, par_lo: float = 20.0,
                             par_hi: float = 220.0) -> pd.DataFrame:
    """Filtra el dataset principal a soberanos en PESOS y clasifica CER/Fija/DL/Dual
    a partir de 'CouponStructure' (dato provisto, no regex sobre ticker).

    Aplica filtros de CALIDAD para que el fit de curva y el carry no se contaminen:
      - solo TradingCcy == 'ARS' (excluye las especies C/D que cotizan en USD)
      - TIR y MD presentes y MD en rango razonable
      - paridad en rango (descarta precios corruptos)
      - opcionalmente descarta bonos legacy/defaulteados (PARP, DICP, CUAP, ...)
    """
    if "Segmento" not in df_norm.columns:
        return pd.DataFrame()
    d = df_norm[df_norm["Segmento"] == "Sovereign"].copy()
    d["Clase"] = d["CouponStructure"].map(COUPON_TO_CLASE).fillna("Otro")

    # Solo especies en pesos (evita AL30C/AL30D y demás que cotizan en dólares)
    if "TradingCcy" in d.columns:
        d = d[d["TradingCcy"].astype(str).str.upper().eq("ARS")]

    # Solo clases operables para el análisis de tasa en pesos
    d = d[d["Clase"].isin(["CER", "Fija", "DL", "Dual"])].copy()

    # Filtros de calidad numérica
    d = d.dropna(subset=["TIR", "MD"])
    d = d[(d["MD"] >= min_md) & (d["MD"] <= max_md)]
    if "Paridad" in d.columns:
        par = pd.to_numeric(d["Paridad"], errors="coerce")
        d = d[par.isna() | ((par >= par_lo) & (par <= par_hi))]

    if drop_legacy:
        tk = d["Ticker"].astype(str).str.upper()
        mask_legacy = tk.str.startswith(LEGACY_PREFIXES)
        d = d[~mask_legacy]

    # Deduplicar por ticker (quedarse con el de mayor paridad = especie más líquida)
    if "Paridad" in d.columns and not d.empty:
        d = (d.sort_values("Paridad", ascending=False)
               .drop_duplicates(subset=["Ticker"], keep="first"))
    return d.reset_index(drop=True)


# =====================================================================
# Nelson-Siegel + cheap/rich
# =====================================================================

def nelson_siegel(md, b0, b1, b2, tau):
    md = np.asarray(md, dtype=float)
    x = md / tau
    with np.errstate(divide="ignore", invalid="ignore"):
        f = np.where(x == 0, 1.0, (1 - np.exp(-x)) / x)
    return b0 + b1 * f + b2 * (f - np.exp(-x))


def fit_ns(md: pd.Series, tir: pd.Series):
    """Ajusta NS con bounds razonables. Devuelve (params, cov) o (None, None)."""
    md, tir = np.asarray(md, float), np.asarray(tir, float)
    if len(md) < 5:
        return None, None
    try:
        p0 = [float(np.nanmean(tir)), -1.0, 1.0, 1.5]
        bounds = ([-50, -80, -80, 0.05], [120, 80, 80, 30])
        params, cov = curve_fit(nelson_siegel, md, tir, p0=p0, bounds=bounds, maxfev=20000)
        return params, cov
    except Exception:
        return None, None


def cheap_rich(df: pd.DataFrame, params) -> pd.DataFrame:
    """Residuo vs curva NS en bps. Positivo = rinde por encima de la curva (barato)."""
    out = df.copy()
    if params is None:
        out["TIR_Curva"] = np.nan
        out["Residuo_bps"] = np.nan
        return out
    out["TIR_Curva"] = nelson_siegel(out["MD"], *params).round(2)
    out["Residuo_bps"] = ((out["TIR"] - out["TIR_Curva"]) * 100).round(0)
    out["Señal"] = np.select(
        [out["Residuo_bps"] >= 50, out["Residuo_bps"] <= -50],
        ["🟢 Barato vs curva", "🔴 Caro vs curva"], default="⚪ En línea",
    )
    return out


def plot_curve_pro(df: pd.DataFrame, title: str, fecha=None, group_col: str | None = None,
                   group_colors: dict | None = None, fit_per_group: bool = False):
    """Scatter TIR vs MD con etiquetas, fit NS y coloreo cheap/rich."""
    dfp = df.dropna(subset=["MD", "TIR"]).copy()
    if dfp.empty:
        return None, None

    fig = go.Figure()
    all_params = {}

    def _hover(d):
        cols = [c for c in ["Paridad", "Convexidad", "Calificacion", "law"] if c in d.columns]
        base = "<b>%{text}</b><br>TIR: %{y:.2f}%<br>MD: %{x:.2f}"
        cd = None
        if cols:
            cd = d[cols].values
            for i, c in enumerate(cols):
                base += f"<br>{c}: %{{customdata[{i}]}}"
        return base + "<extra></extra>", cd

    groups = [(None, dfp)] if not group_col else list(dfp.groupby(group_col))
    for gname, dg in groups:
        color = (group_colors or {}).get(gname, COLORS["primary"])
        hv, cd = _hover(dg)
        fig.add_trace(go.Scatter(
            x=dg["MD"], y=dg["TIR"], mode="markers+text",
            text=dg["Ticker"], textposition="top center",
            textfont=dict(size=10, color="#334155"),
            marker=dict(size=11, color=color, line=dict(width=1.2, color="white"), opacity=0.9),
            name=str(gname) if gname else "Instrumentos",
            customdata=cd, hovertemplate=hv,
        ))
        if fit_per_group:
            p, _ = fit_ns(dg["MD"], dg["TIR"])
            all_params[gname] = p
            if p is not None and len(dg) >= 5:
                dom = np.linspace(dg["MD"].min(), dg["MD"].max(), 120)
                fig.add_trace(go.Scatter(x=dom, y=nelson_siegel(dom, *p), mode="lines",
                                         line=dict(dash="dash", width=1.6, color=color),
                                         name=f"NS {gname}", hoverinfo="skip", opacity=0.6))

    if not fit_per_group:
        p, cov = fit_ns(dfp["MD"], dfp["TIR"])
        all_params["global"] = p
        if p is not None:
            dom = np.linspace(dfp["MD"].min(), dfp["MD"].max(), 150)
            fitted = nelson_siegel(dom, *p)
            fig.add_trace(go.Scatter(x=dom, y=fitted, mode="lines",
                                     line=dict(dash="dash", width=2, color=COLORS["fit"]),
                                     name="Curva Nelson-Siegel"))
            fig.add_trace(go.Scatter(x=np.concatenate([dom, dom[::-1]]),
                                     y=np.concatenate([fitted + 0.5, (fitted - 0.5)[::-1]]),
                                     fill="toself", fillcolor="rgba(100,116,139,0.10)",
                                     line=dict(width=0), name="±50 bps", hoverinfo="skip"))

    fig.update_layout(title=title, height=520, **PLOTLY_LAYOUT)
    style_axes(fig, "Modified Duration (años)", "TIR (%)")
    add_source_watermark(fig, fecha)
    return fig, all_params


# =====================================================================
# Spreads por legislación (AL/AE vs GD) + z-score
# =====================================================================

def classify_bono(ticker: str) -> str:
    if ticker in BONOS_GD:
        return "GD"
    if ticker in BONOS_AL:
        return "AL"
    if ticker in BOPREAL:
        return "BOPREAL"
    return "Otro"


def spread_series(df_norm: pd.DataFrame) -> pd.DataFrame:
    todos = BONOS_GD + BONOS_AL
    d = df_norm[df_norm["Ticker"].isin(todos)].dropna(subset=["TIR"]).copy()
    if d.empty or "Date" not in d.columns:
        return pd.DataFrame()
    d["Numero"] = d["Ticker"].str.extract(r"(\d+)")
    d["Grupo"] = np.where(d["Ticker"].str.startswith("GD"), "GD", "AL")
    piv = d.pivot_table(index=["Date", "Numero"], columns="Grupo", values="TIR", aggfunc="first").reset_index()
    if not {"AL", "GD"}.issubset(piv.columns):
        return pd.DataFrame()
    piv = piv.dropna(subset=["AL", "GD"])
    piv["Spread"] = piv["AL"] - piv["GD"]
    return piv


def spread_stats(spreads: pd.DataFrame) -> pd.DataFrame:
    if spreads.empty:
        return pd.DataFrame()
    rows = []
    for num, g in spreads.groupby("Numero"):
        g = g.sort_values("Date").tail(252)
        if g.empty:
            continue
        cur, mu, sd = g["Spread"].iloc[-1], g["Spread"].mean(), g["Spread"].std()
        z = (cur - mu) / sd if sd and sd > 0 else np.nan
        rows.append({"Numero": int(num), "Spread_Actual": round(cur, 2),
                     "Spread_Prom_252": round(mu, 2), "Desvio_252": round(sd, 2) if pd.notna(sd) else np.nan,
                     "Z_Score": round(z, 2) if pd.notna(z) else np.nan})
    out = pd.DataFrame(rows).sort_values("Numero")
    if not out.empty:
        out["Señal"] = np.select(
            [out["Z_Score"] >= 1.5, out["Z_Score"] <= -1.5],
            ["🟢 Spread caro → favorece ley local (AL)", "🔵 Spread comprimido → favorece ley NY (GD)"],
            default="⚪ Neutral",
        )
    return out


def plot_spread_history(spreads: pd.DataFrame):
    if spreads.empty:
        return None
    fig = go.Figure()
    palette = ["#2563eb", "#dc2626", "#059669", "#7c3aed", "#ea580c", "#0891b2"]
    for i, (num, g) in enumerate(spreads.groupby("Numero")):
        g = g.sort_values("Date").tail(252)
        fig.add_trace(go.Scatter(x=g["Date"], y=g["Spread"], mode="lines",
                                 name=f"20{num}" if len(str(num)) == 2 else str(num),
                                 line=dict(width=1.8, color=palette[i % len(palette)])))
    fig.add_hline(y=0, line_dash="dot", line_color="#94a3b8")
    fig.update_layout(title="Spread por legislación (AL/AE − GD) · últimas 252 ruedas",
                      height=440, **PLOTLY_LAYOUT)
    style_axes(fig, "Fecha", "Spread (puntos de TIR)")
    add_source_watermark(fig)
    return fig


# =====================================================================
# Breakevens (CER / DL vs tasa fija)
# =====================================================================

def compute_breakevens(df_latest: pd.DataFrame, kind: str = "CER") -> pd.DataFrame:
    """
    Breakeven = (1 + TIR_fija) / (1 + TIR_real_o_DL) − 1, interpolando la curva de
    tasa fija (fit NS o interpolación lineal) en la MD de cada bono CER / DL.
      * CER → inflación breakeven anual implícita
      * DL  → devaluación breakeven anual implícita
    """
    fija = df_latest[df_latest["Clase"] == "Fija"].dropna(subset=["MD", "TIR"])
    target = df_latest[df_latest["Clase"] == kind].dropna(subset=["MD", "TIR"])
    if len(fija) < 2 or target.empty:
        return pd.DataFrame()

    params, _ = fit_ns(fija["MD"], fija["TIR"])

    def tasa_fija_en(md):
        lo, hi = fija["MD"].min(), fija["MD"].max()
        if params is not None and lo <= md <= hi:
            return float(nelson_siegel(np.array([md]), *params)[0])
        f = fija.sort_values("MD")
        return float(np.interp(md, f["MD"], f["TIR"]))

    rows = []
    lo, hi = fija["MD"].min(), fija["MD"].max()
    for _, r in target.iterrows():
        md = float(r["MD"])
        tn = tasa_fija_en(md)
        be = ((1 + tn / 100) / (1 + float(r["TIR"]) / 100) - 1) * 100
        rows.append({"Ticker": r["Ticker"], "MD": round(md, 2),
                     "TIR_" + ("Real" if kind == "CER" else "DL"): round(float(r["TIR"]), 2),
                     "TIR_Fija_Interp": round(tn, 2),
                     "Breakeven_%anual": round(be, 2),
                     "Extrapolado": md < lo or md > hi})
    return pd.DataFrame(rows).sort_values("MD")


def plot_breakeven(df_be: pd.DataFrame, title: str, color: str, ref_value: float | None = None,
                   ref_label: str | None = None):
    if df_be.empty:
        return None
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df_be["MD"], y=df_be["Breakeven_%anual"], mode="markers+lines+text",
        text=df_be["Ticker"], textposition="top center", textfont=dict(size=10),
        line=dict(width=2, color=color), marker=dict(size=10, color=color, line=dict(width=1, color="white")),
        name="Breakeven implícito",
        hovertemplate="<b>%{text}</b><br>MD: %{x:.2f}<br>Breakeven: %{y:.2f}% anual<extra></extra>",
    ))
    if ref_value is not None:
        fig.add_hline(y=ref_value, line_dash="dash", line_color=COLORS["accent"],
                      annotation_text=ref_label or f"Referencia: {ref_value:.1f}%",
                      annotation_font_color=COLORS["accent"])
    fig.update_layout(title=title, height=460, **PLOTLY_LAYOUT)
    style_axes(fig, "Modified Duration (años)", "Breakeven (% anual)")
    add_source_watermark(fig)
    return fig


# =====================================================================
# REM (dataset 44033) — inflación esperada a 12 meses
# =====================================================================

def get_rem_inflacion_12m(df_rem_raw: pd.DataFrame):
    """Devuelve (mediana_%, fecha_relevamiento) para la inflación i.a. esperada a
    'Próx. 12 meses' según el IPC nivel general del REM. None, None si no hay datos."""
    needed = {"Variable", "Período", "Mediana", "Date"}
    if df_rem_raw is None or df_rem_raw.empty or not needed.issubset(df_rem_raw.columns):
        return None, None
    d = df_rem_raw.copy()
    d = d[(d["Variable"] == "IPC nivel general") & (d["Período"] == "Próx. 12 meses")]
    if d.empty:
        return None, None
    d["Date"] = pd.to_datetime(d["Date"], errors="coerce")
    d = d.dropna(subset=["Date"]).sort_values("Date")
    if d.empty:
        return None, None
    last = d.iloc[-1]
    val = pd.to_numeric(last["Mediana"], errors="coerce")
    if pd.isna(val):
        return None, None
    return float(val), last["Date"]


# =====================================================================
# Futuros ROFEX (dataset 5361) — curva implícita de devaluación
# =====================================================================

def get_futuros_curva(df_fut_raw: pd.DataFrame):
    """Devuelve (df_curva, spot, fecha, diag).

    El dataset 5361 ('Estimated Implied Curve') tiene columnas Spot y '<N> days'.
    OJO: no siempre los tenors están en la misma escala que el Spot (a veces son
    precio $/USD, a veces vienen como índice). Esta versión:
      - toma la ÚLTIMA fila con Spot válido,
      - por cada tenor usa el último valor NO NULO y > 0 (no fuerza la última fila,
        que suele tener tenors largos sin cotizar → causaba el bug de -100%),
      - valida que el precio del futuro sea plausible vs spot (0.8x a 5x): si no,
        descarta ese tenor en vez de generar devaluaciones imposibles,
      - devuelve un dict 'diag' para auditoría.
    df_curva: Días, Precio, Deval_periodo_%, Deval_anualizada_%.
    """
    diag = {"ok": False, "motivo": ""}
    if df_fut_raw is None or df_fut_raw.empty or "Spot" not in df_fut_raw.columns:
        diag["motivo"] = "sin columna Spot"
        return pd.DataFrame(), None, None, diag
    d = df_fut_raw.copy()
    if "Date" in d.columns:
        d["Date"] = pd.to_datetime(d["Date"], errors="coerce")
        d = d.dropna(subset=["Date"]).sort_values("Date")
    if d.empty:
        diag["motivo"] = "sin filas con fecha"
        return pd.DataFrame(), None, None, diag

    # Último spot válido
    spot_series = pd.to_numeric(d["Spot"], errors="coerce").dropna()
    if spot_series.empty or spot_series.iloc[-1] <= 0:
        diag["motivo"] = "spot inválido"
        return pd.DataFrame(), None, d["Date"].iloc[-1] if "Date" in d else None, diag
    spot = float(spot_series.iloc[-1])
    fecha = d.loc[spot_series.index[-1], "Date"] if "Date" in d.columns else None

    # Tenors "<N> days" SIN sufijos derivados (change / running_av / mom)
    tenor_cols = []
    for c in d.columns:
        if not isinstance(c, str):
            continue
        cs = c.strip()
        if cs.lower().endswith(" days") and cs.split(" ")[0].isdigit():
            tenor_cols.append(c)

    rows = []
    for c in tenor_cols:
        serie = pd.to_numeric(d[c], errors="coerce").dropna()
        serie = serie[serie > 0]
        if serie.empty:
            continue
        px = float(serie.iloc[-1])
        days = int(c.strip().split(" ")[0])
        ratio = px / spot
        # Validación de plausibilidad: un futuro de dólar cotiza por encima del spot
        # pero no 5x en <1 año. Si el ratio es absurdo, el tenor está en otra escala.
        if not (0.8 <= ratio <= 5.0):
            continue
        deval_periodo = (ratio - 1) * 100
        deval_anual = (ratio ** (365.0 / days) - 1) * 100
        rows.append({"Días": days, "Precio": round(px, 2),
                     "Deval_periodo_%": round(deval_periodo, 2),
                     "Deval_anualizada_%": round(deval_anual, 2)})

    if not rows:
        diag["ok"] = False
        diag["tenors_validos"] = 0
        diag["tenors_totales"] = len(tenor_cols)
        diag["motivo"] = ("ningún tenor pasó la validación de plausibilidad vs spot "
                          "(precios en escala distinta al spot, o sin cotización) — "
                          "se usará devaluación implícita de los bonos DL")
        return pd.DataFrame(), spot, fecha, diag

    df_curva = pd.DataFrame(rows).sort_values("Días").reset_index(drop=True)
    diag["ok"] = True
    diag["tenors_validos"] = len(df_curva)
    diag["tenors_totales"] = len(tenor_cols)
    return df_curva, spot, fecha, diag


def deval_implicita_desde_bonos(fija_df: pd.DataFrame, dl_df: pd.DataFrame,
                                params_fija=None, horizon_days: float = 365) -> float | None:
    """Fallback robusto: devaluación anualizada implícita estimada desde los propios
    bonos, como la que iguala fija vs DL a ~1 año de duration. Es el breakeven de
    devaluación de mercado (BE = (1+TIR_fija)/(1+TIR_DL)-1) evaluado en la duration
    más cercana a 'horizon_days'. No depende del dataset de futuros."""
    if fija_df is None or dl_df is None or fija_df.empty or dl_df.empty:
        return None
    target_md = horizon_days / 365.0
    dl_sorted = dl_df.dropna(subset=["MD", "TIR"]).copy()
    if dl_sorted.empty:
        return None
    dl_sorted["dist"] = (dl_sorted["MD"] - target_md).abs()
    r = dl_sorted.sort_values("dist").iloc[0]
    md = float(r["MD"]); tir_dl = float(r["TIR"])

    lo, hi = fija_df["MD"].min(), fija_df["MD"].max()
    if params_fija is not None and lo <= md <= hi:
        tir_fija = float(nelson_siegel(np.array([md]), *params_fija)[0])
    else:
        f = fija_df.sort_values("MD")
        tir_fija = float(np.interp(md, f["MD"], f["TIR"]))
    be = ((1 + tir_fija / 100) / (1 + tir_dl / 100) - 1) * 100
    return round(be, 1)


def interp_futuros_annual(df_curva: pd.DataFrame, days: float):
    """Interpola (o extrapola plano) la devaluación anualizada implícita a 'days'."""
    if df_curva is None or df_curva.empty:
        return np.nan
    dom = df_curva["Días"].values.astype(float)
    y = df_curva["Deval_anualizada_%"].values.astype(float)
    if len(dom) == 1:
        return float(y[0])
    if days <= dom.max():
        return float(np.interp(days, dom, y))
    return float(y[-1])


# =====================================================================
# Carry Trade: breakeven de mercado vs expectativa (REM / ROFEX)
# =====================================================================

def carry_scenario(df_target: pd.DataFrame, fija: pd.DataFrame, kind: str,
                   escenario_anual: float | None, params_fija=None) -> pd.DataFrame:
    """
    Motor de carry por ESCENARIO. Para cada bono de cobertura (CER o DL) calcula,
    a igual duration, el retorno total esperado de la cobertura vs. el de la tasa
    fija comparable, bajo un supuesto de inflación/devaluación anual dado.

    Definiciones (todo en % anual efectivo salvo aclaración):
      - TIR_cobertura = TIR del bono en su moneda de indexación (real para CER, en
        USD para DL). Es lo que rinde el bono POR ENCIMA del ajuste.
      - Retorno_esperado_cobertura ≈ (1+TIR_cob/100)*(1+escenario/100) - 1
        (rinde el ajuste + su spread real/USD).
      - Retorno_fija = tasa fija nominal interpolada a la misma duration.
      - Excess_pp = Retorno_esperado_cobertura - Retorno_fija.
          Excess > 0  → conviene la COBERTURA si el escenario se cumple.
          Excess < 0  → conviene la tasa FIJA si el escenario se cumple.
      - Breakeven = escenario de indexación que iguala ambas patas (Excess = 0):
          BE = (1+TIR_fija)/(1+TIR_cob) - 1.  Es el punto de indiferencia.
      - Margen_vs_BE_pp = escenario_usuario - Breakeven. Cuánta inflación/deval
        de más (o de menos) estás asumiendo respecto del punto de equilibrio.
    """
    out = df_target.dropna(subset=["MD", "TIR"]).copy()
    if out.empty or len(fija) < 2:
        return pd.DataFrame()

    def tasa_fija_en(md):
        lo, hi = fija["MD"].min(), fija["MD"].max()
        if params_fija is not None and lo <= md <= hi:
            return float(nelson_siegel(np.array([md]), *params_fija)[0])
        f = fija.sort_values("MD")
        return float(np.interp(md, f["MD"], f["TIR"]))

    col_tir = "TIR_Real" if kind == "CER" else "TIR_DL"
    rows = []
    for _, r in out.iterrows():
        md = float(r["MD"])
        tir_cob = float(r["TIR"])
        tir_fija = tasa_fija_en(md)
        be = ((1 + tir_fija / 100) / (1 + tir_cob / 100) - 1) * 100
        row = {
            "Ticker": r["Ticker"], "MD": round(md, 2),
            col_tir: round(tir_cob, 2),
            "TIR_Fija_interp": round(tir_fija, 2),
            "Breakeven_%anual": round(be, 2),
        }
        if escenario_anual is not None:
            ret_cob = ((1 + tir_cob / 100) * (1 + escenario_anual / 100) - 1) * 100
            excess = ret_cob - tir_fija
            row.update({
                "Escenario_%anual": round(escenario_anual, 2),
                "Ret_cobertura_%": round(ret_cob, 2),
                "Ret_fija_%": round(tir_fija, 2),
                "Excess_pp": round(excess, 2),
                "Margen_vs_BE_pp": round(escenario_anual - be, 2),
            })
        rows.append(row)

    res = pd.DataFrame(rows).sort_values("MD").reset_index(drop=True)
    if "Excess_pp" in res.columns:
        res["Recomendación"] = np.select(
            [res["Excess_pp"] >= 1.0, res["Excess_pp"] <= -1.0],
            [f"🟢 Cobertura {kind}", "🔵 Tasa Fija"], default="⚪ Indiferente",
        )
    return res


# =====================================================================
# Sensibilidad: mapa de calor excess return por instrumento y escenario
# =====================================================================

def sensitivity_matrix(bono_tir: float, tir_fija_interp: float, escenarios: np.ndarray) -> np.ndarray:
    """Vector de excess return (pp) de una cobertura vs su fija comparable, para
    una grilla de escenarios de indexación (inflación o devaluación).
      excess(e) = [(1+TIR_cob)*(1+e) - 1] - TIR_fija   (todo en %)
    """
    ret_cob = ((1 + bono_tir / 100) * (1 + escenarios / 100) - 1) * 100
    return ret_cob - tir_fija_interp


def build_heatmap(df_carry: pd.DataFrame, kind: str, esc_min: float, esc_max: float,
                  n: int = 25, escenario_marker: float | None = None):
    """Heatmap: filas = instrumentos (ordenados por MD), columnas = escenario de
    inflación/devaluación anual, color = excess return (pp) de la cobertura vs fija.
    Verde = conviene la cobertura; rojo = conviene la fija. La línea vertical marca
    el escenario elegido; la 'X' por fila marca el breakeven de cada bono."""
    if df_carry.empty or "TIR_" + ("Real" if kind == "CER" else "DL") not in df_carry.columns:
        return None
    col_tir = "TIR_Real" if kind == "CER" else "TIR_DL"
    d = df_carry.dropna(subset=[col_tir, "TIR_Fija_interp", "MD"]).sort_values("MD")
    if d.empty:
        return None

    escenarios = np.linspace(esc_min, esc_max, n)
    z = np.array([sensitivity_matrix(float(r[col_tir]), float(r["TIR_Fija_interp"]), escenarios)
                  for _, r in d.iterrows()])
    y_labels = [f"{r['Ticker']} · MD {r['MD']:.2f}" for _, r in d.iterrows()]

    zmax = np.nanmax(np.abs(z)) if z.size else 1
    fig = go.Figure(data=go.Heatmap(
        z=z, x=escenarios, y=y_labels,
        colorscale=[[0, "#b91c1c"], [0.5, "#f8fafc"], [1, "#15803d"]],
        zmid=0, zmin=-zmax, zmax=zmax,
        colorbar=dict(title="Excess<br>(pp)", thickness=14),
        hovertemplate=("<b>%{y}</b><br>Escenario: %{x:.1f}%<br>"
                       "Excess cobertura vs fija: %{z:+.1f} pp<extra></extra>"),
    ))

    # marca de breakeven por fila (donde excess cruza 0)
    be_x = d["Breakeven_%anual"].values.astype(float)
    fig.add_trace(go.Scatter(
        x=be_x, y=y_labels, mode="markers",
        marker=dict(symbol="x", size=9, color="#0f172a", line=dict(width=1, color="white")),
        name="Breakeven", hovertemplate="Breakeven: %{x:.1f}%<extra></extra>",
    ))

    if escenario_marker is not None:
        fig.add_vline(x=escenario_marker, line_dash="dash", line_color=COLORS["accent"],
                      annotation_text=f"Escenario: {escenario_marker:.1f}%",
                      annotation_font_color=COLORS["accent"])

    etiqueta = "inflación" if kind == "CER" else "devaluación"
    fig.update_layout(title=f"Sensibilidad: excess return de {kind} vs Fija según {etiqueta} anual",
                      height=max(300, 44 * len(y_labels) + 140), **PLOTLY_LAYOUT)
    style_axes(fig, f"{etiqueta.capitalize()} anual asumida (%)", "")
    fig.update_yaxes(autorange="reversed")
    add_source_watermark(fig)
    return fig


def split_curve_outliers(df: pd.DataFrame, min_md: float = 0.3,
                         tir_lo: float = -10.0, tir_hi: float = 200.0):
    """Separa el universo en (limpios, sospechosos) para el fit de curva.

    Un instrumento va a 'sospechosos' si:
      - MD < min_md (típicamente < 0.3 años: a días de vencer la TIR anualizada se
        vuelve numéricamente inestable — el interés corrido domina el precio), o
      - TIR fuera de un rango sensato para pesos/USD (tasa real muy negativa o
        nominal explosiva) que delata precio o cupón mal cargado.

    Los 'limpios' alimentan el fit Nelson-Siegel y los breakevens; los 'sospechosos'
    se muestran aparte con un warning, no se ocultan (el analista debe verlos)."""
    if df is None or df.empty:
        return df, pd.DataFrame()
    d = df.copy()
    tir = pd.to_numeric(d["TIR"], errors="coerce")
    md = pd.to_numeric(d["MD"], errors="coerce")
    ok = (md >= min_md) & (tir >= tir_lo) & (tir <= tir_hi)
    return d[ok].reset_index(drop=True), d[~ok].reset_index(drop=True)


def drop_short_outliers(df: pd.DataFrame, min_md: float = 0.3,
                        tir_lo: float = -10.0, tir_hi: float = 200.0) -> pd.DataFrame:
    """Wrapper retrocompatible: devuelve solo los limpios."""
    limpios, _ = split_curve_outliers(df, min_md, tir_lo, tir_hi)
    return limpios


# =====================================================================
# Informe Excel — snapshot con las tablas e insights clave
# =====================================================================

_XLS_FONT = "Arial"
_XLS_HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
_XLS_HEADER_FONT = Font(name=_XLS_FONT, bold=True, color="FFFFFF", size=10)
_XLS_TITLE_FONT = Font(name=_XLS_FONT, bold=True, size=14, color="1F2A44")
_XLS_SUBTITLE_FONT = Font(name=_XLS_FONT, size=10, italic=True, color="64748B")
_XLS_LABEL_FONT = Font(name=_XLS_FONT, bold=True, size=10, color="1F2A44")
_XLS_BODY_FONT = Font(name=_XLS_FONT, size=10)
_XLS_NOTE_FONT = Font(name=_XLS_FONT, size=9, italic=True, color="64748B")
_XLS_BORDER = Border(*(Side(style="thin", color="D1D5DB") for _ in range(4)))
_XLS_ALERT_FILL = PatternFill("solid", fgColor="FEF3C7")

# El informe es un SNAPSHOT (valores, no fórmulas): reproduce el estado de la app
# en el momento de la descarga, no un modelo vivo que se recalcule en Excel.
# Los textos de la UI ya son descriptivos ("🟢 Barato vs curva", "🔸 Baja"); el emoji es
# redundante en un documento formal, así que simplemente se retira (no se reemplaza por
# una etiqueta) para no duplicar la información en la celda.
_EMOJI_STRIP = ["🟢", "🔴", "🔵", "⚪", "✅", "🔸", "🎯", "📌", "⚠️", "ℹ️", "⚙️", "**"]


def _xls_clean(val):
    """Convierte emojis/markdown de la UI a texto plano apto para un documento formal."""
    if isinstance(val, str):
        s = val
        for tok in _EMOJI_STRIP:
            s = s.replace(tok, "")
        return " ".join(s.split())
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, float) and pd.isna(val):
        return None
    return val


def _xls_write_table(ws, df: pd.DataFrame, start_row: int, title: str | None = None) -> int:
    """Escribe un DataFrame formateado (encabezado, bordes, ancho de columna) y
    devuelve la fila siguiente libre."""
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = _XLS_TITLE_FONT
        r += 2
    if df is None or df.empty:
        ws.cell(row=r, column=1, value="Sin datos disponibles para esta tabla.").font = _XLS_SUBTITLE_FONT
        return r + 2

    cols = list(df.columns)
    for j, col in enumerate(cols, start=1):
        c = ws.cell(row=r, column=j, value=str(col))
        c.font = _XLS_HEADER_FONT
        c.fill = _XLS_HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _XLS_BORDER
    header_row = r
    r += 1

    for _, row in df.iterrows():
        for j, col in enumerate(cols, start=1):
            val = _xls_clean(row[col])
            c = ws.cell(row=r, column=j, value=val)
            c.font = _XLS_BODY_FONT
            c.border = _XLS_BORDER
            if isinstance(val, (int, float)):
                c.number_format = "0.00"
                c.alignment = Alignment(horizontal="right")
            else:
                c.alignment = Alignment(horizontal="left")
        r += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    for j, col in enumerate(cols, start=1):
        largo = [len(str(col))]
        for v in df[col].tolist():
            vv = _xls_clean(v)
            largo.append(len(str(vv)) if vv is not None else 0)
        ws.column_dimensions[get_column_letter(j)].width = min(max(10, max(largo) + 2), 42)
    return r + 2


def _xls_write_kv(ws, row: int, label: str, value, note: str | None = None) -> int:
    """Escribe un par etiqueta/valor (para el resumen ejecutivo) y devuelve la fila siguiente."""
    ws.cell(row=row, column=1, value=label).font = _XLS_LABEL_FONT
    c = ws.cell(row=row, column=2, value=_xls_clean(value))
    c.font = _XLS_BODY_FONT
    if note:
        ws.cell(row=row, column=3, value=note).font = _XLS_NOTE_FONT
    return row + 1


# ---------------------------------------------------------------------------
# Gráficos nativos de Excel (editables) para la hoja Resumen.
# Los datos de cada gráfico se escriben en una hoja auxiliar OCULTA (el chart de
# Excel necesita referenciar celdas reales, no puede tomar un DataFrame directo).
# ---------------------------------------------------------------------------

def _hidden_data_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.sheet_state = "hidden"
    return ws


def _write_columns(ws, headers: list, columns: list, start_row: int = 1, start_col: int = 1) -> dict:
    """Escribe columnas paralelas (una lista de listas, todas la misma cantidad de
    filas o menos) en 'ws' a partir de (start_row, start_col). Devuelve un dict
    {header: (col_idx, first_data_row, last_data_row)} para armar Reference() después."""
    info = {}
    for j, (h, coldata) in enumerate(zip(headers, columns)):
        col = start_col + j
        ws.cell(row=start_row, column=col, value=h).font = _XLS_HEADER_FONT
        n = 0
        for i, v in enumerate(coldata):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            ws.cell(row=start_row + 1 + i, column=col, value=float(v) if isinstance(v, (int, float)) else v)
            n += 1
        info[h] = (col, start_row + 1, start_row + max(n, 1))
    return info


def _scatter_chart(title: str, x_title: str, y_title: str, height: float = 8.5, width: float = 16.0) -> ScatterChart:
    ch = ScatterChart()
    ch.title = title
    ch.style = 2
    ch.x_axis.title = x_title
    ch.y_axis.title = y_title
    # CRÍTICO: en un ScatterChart ambos ejes son NumericAxis y openpyxl no fija
    # axPos por separado — sin esto, Excel real (no LibreOffice) marca el archivo
    # como corrupto porque quedan los dos ejes como "izquierda" (axPos='l').
    ch.x_axis.axPos = "b"
    ch.y_axis.axPos = "l"
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.x_axis.crosses = "min"
    ch.y_axis.crosses = "min"
    ch.height = height
    ch.width = width
    return ch


def _add_xy_series(chart: ScatterChart, data_ws, x_info: tuple, y_info: tuple, name: str,
                   markers_only: bool = False, line_only: bool = False, color: str | None = None):
    """x_info/y_info = (col_idx, first_row, last_row) de _write_columns."""
    xcol, xr0, xr1 = x_info
    ycol, yr0, yr1 = y_info
    last = min(xr1, yr1)
    xref = Reference(data_ws, min_col=xcol, min_row=xr0, max_row=last)
    yref = Reference(data_ws, min_col=ycol, min_row=yr0, max_row=last)
    s = Series(yref, xref, title=name)
    if markers_only:
        s.marker = Marker(symbol="circle", size=6)
        s.graphicalProperties.line.noFill = True
    if line_only:
        s.marker = Marker(symbol="none")
        s.smooth = False
    if color:
        s.marker.graphicalProperties.solidFill = color
        s.graphicalProperties.line.solidFill = color
    chart.series.append(s)


def _pivot_spread_history(spreads: pd.DataFrame, max_rows: int = 120) -> pd.DataFrame:
    """De spread_series() (Date, Numero, Spread largo) arma una tabla ancha
    Date + una columna por año de vencimiento, para un LineChart nativo."""
    if spreads is None or spreads.empty:
        return pd.DataFrame()
    d = spreads.copy().sort_values("Date")
    piv = d.pivot_table(index="Date", columns="Numero", values="Spread", aggfunc="first")
    piv = piv.tail(max_rows).reset_index()
    piv.columns = ["Date"] + [f"20{c}" if len(str(c)) == 2 else str(c) for c in piv.columns[1:]]
    return piv


def _add_sensitivity_table(ws, df_carry: pd.DataFrame, kind: str, start_row: int, start_col: int = 6,
                           n_cols: int = 9) -> int:
    """Tabla de sensibilidad (instrumento x escenario) con formato condicional
    rojo→blanco→verde, equivalente al heatmap de la app pero nativo de Excel.
    Devuelve la fila siguiente libre."""
    col_tir = "TIR_Real" if kind == "CER" else "TIR_DL"
    if df_carry is None or df_carry.empty or col_tir not in df_carry.columns:
        ws.cell(row=start_row, column=start_col,
               value=f"Sensibilidad {kind}: sin datos.").font = _XLS_SUBTITLE_FONT
        return start_row + 2

    d = df_carry.dropna(subset=[col_tir, "TIR_Fija_interp", "MD"]).sort_values("MD")
    if d.empty:
        ws.cell(row=start_row, column=start_col,
               value=f"Sensibilidad {kind}: sin datos.").font = _XLS_SUBTITLE_FONT
        return start_row + 2

    lo = float(d["Breakeven_%anual"].min()) - 8
    hi = float(d["Breakeven_%anual"].max()) + 12
    escenarios = np.linspace(max(lo, -5), hi, n_cols)

    etiqueta = "Inflación" if kind == "CER" else "Devaluación"
    ws.cell(row=start_row, column=start_col,
           value=f"Sensibilidad {kind} vs Fija — excess return (pp) según {etiqueta.lower()} anual asumida").font = _XLS_TITLE_FONT
    r = start_row + 2

    headers = ["Instrumento", "MD", "Breakeven %"] + [f"{e:.0f}%" for e in escenarios]
    for j, h in enumerate(headers):
        c = ws.cell(row=r, column=start_col + j, value=h)
        c.font = _XLS_HEADER_FONT
        c.fill = _XLS_HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = _XLS_BORDER
    header_row = r
    r += 1

    for _, row in d.iterrows():
        ws.cell(row=r, column=start_col, value=row["Ticker"]).font = _XLS_BODY_FONT
        ws.cell(row=r, column=start_col + 1, value=round(float(row["MD"]), 2)).font = _XLS_BODY_FONT
        ws.cell(row=r, column=start_col + 2, value=round(float(row["Breakeven_%anual"]), 1)).font = _XLS_BODY_FONT
        vals = sensitivity_matrix(float(row[col_tir]), float(row["TIR_Fija_interp"]), escenarios)
        for j, v in enumerate(vals):
            c = ws.cell(row=r, column=start_col + 3 + j, value=round(float(v), 1))
            c.font = _XLS_BODY_FONT
            c.border = _XLS_BORDER
            c.alignment = Alignment(horizontal="center")
        r += 1
    last_row = r - 1

    rng = (f"{get_column_letter(start_col + 3)}{header_row + 1}:"
           f"{get_column_letter(start_col + 2 + n_cols)}{last_row}")
    rule = ColorScaleRule(start_type="min", start_color="B91C1C",
                         mid_type="percentile", mid_value=50, mid_color="F8FAFC",
                         end_type="max", end_color="15803D")
    ws.conditional_formatting.add(rng, rule)

    ws.cell(row=r, column=start_col, value=(
        "Verde = conviene la cobertura bajo ese escenario. Rojo = conviene tasa fija. "
        "La columna 'Breakeven %' es el punto exacto de indiferencia de cada instrumento."
    )).font = _XLS_NOTE_FONT
    return r + 2


def build_excel_report(ctx: dict) -> bytes:
    """Arma el informe de Renta Fija en Excel (bytes, listo para st.download_button).

    'ctx' es un diccionario con las tablas y valores ya calculados en la app
    (ver el punto de armado al final del script). No usa fórmulas: es la foto de
    un momento dado, igual que lo que se ve en pantalla al momento de exportar.
    """
    wb = Workbook()

    # ---------- Hoja 1: Resumen ejecutivo ----------
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 55

    ws.cell(row=1, column=1, value="Renta Fija Argentina — Informe PRO").font = _XLS_TITLE_FONT
    ws.cell(row=2, column=1, value=(
        f"Generado {datetime.now().strftime('%d/%m/%Y %H:%M')} · Datos al "
        f"{ctx.get('fecha_datos', '—')} · Herramienta de análisis, no recomendación de inversión."
    )).font = _XLS_SUBTITLE_FONT

    r = 4
    ws.cell(row=r, column=1, value="Escenario macro utilizado").font = _XLS_LABEL_FONT
    r += 1
    r = _xls_write_kv(ws, r, "Inflación esperada (% i.a.)", ctx.get("infl_esc"),
                      f"Fuente: REM {ctx.get('rem_val_txt','—')}" if ctx.get("infl_es_rem") else "Escenario propio del usuario")
    r = _xls_write_kv(ws, r, "Devaluación esperada (% i.a.)", ctx.get("deval_esc"),
                      f"Fuente: {ctx.get('deval_source','—')}")
    r += 1

    ws.cell(row=r, column=1, value="Tablero de decisión").font = _XLS_LABEL_FONT
    r += 1
    r = _xls_write_kv(ws, r, "Breakeven inflación (tramo corto MD≤1,5)", ctx.get("be_cer_short"),
                      ctx.get("delta_cer_txt"))
    r = _xls_write_kv(ws, r, "Mejor CER bajo el escenario", ctx.get("best_cer_txt"))
    r = _xls_write_kv(ws, r, "Breakeven devaluación (tramo corto)", ctx.get("be_dl_short"),
                      ctx.get("delta_dl_txt"))
    r = _xls_write_kv(ws, r, "Señal dominante (CER vs Fija)", ctx.get("senal_dominante"))
    r += 1

    if ctx.get("lectura"):
        ws.cell(row=r, column=1, value="Lectura").font = _XLS_LABEL_FONT
        r += 1
        cell = ws.cell(row=r, column=1, value=_xls_clean(ctx["lectura"]))
        cell.font = _XLS_BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 45
        r += 2

    if ctx.get("bullets"):
        ws.cell(row=r, column=1, value="Insights automáticos").font = _XLS_LABEL_FONT
        r += 1
        for b in ctx["bullets"]:
            cell = ws.cell(row=r, column=1, value=f"• {_xls_clean(b)}")
            cell.font = _XLS_BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    if ctx.get("caveats"):
        ws.cell(row=r, column=1, value="Caveats de datos (revisar antes de operar)").font = Font(
            name=_XLS_FONT, bold=True, size=10, color="92400E")
        r += 1
        for cv in ctx["caveats"]:
            cell = ws.cell(row=r, column=1, value=f"• {_xls_clean(cv)}")
            cell.font = _XLS_BODY_FONT
            cell.fill = _XLS_ALERT_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.row_dimensions[r].height = 30
            r += 1
        r += 1

    ws.cell(row=r, column=1, value=(
        "Supuestos: mismo horizonte de tenencia, sin reinversión de cupón ni costos de transacción; "
        "TIR anual efectiva. El breakeven es un punto de indiferencia, no una predicción."
    )).font = _XLS_NOTE_FONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    # ---------- Gráficos nativos (columna F en adelante, no chocan con el texto) ----------
    ws.cell(row=1, column=6, value="Gráficos").font = _XLS_TITLE_FONT
    data_ws = _hidden_data_sheet(wb, "_ChartData")
    chart_anchor_row = 2

    # 1) Curva de ONs: TIR observada (puntos) + TIR teórica de la curva NS (línea)
    try:
        df_ons_c = ctx.get("df_ons")
        if df_ons_c is not None and not df_ons_c.empty and {"MD", "TIR"}.issubset(df_ons_c.columns):
            d = df_ons_c.dropna(subset=["MD", "TIR"]).sort_values("MD")
            info = _write_columns(data_ws, ["ONs_MD", "ONs_TIR", "ONs_TIRCurva"],
                                  [d["MD"].tolist(), d["TIR"].tolist(),
                                   d["TIR_Curva"].tolist() if "TIR_Curva" in d.columns else []],
                                  start_row=1, start_col=1)
            ch = _scatter_chart("Curva de ONs — TIR vs MD", "Modified Duration (años)", "TIR (%)")
            _add_xy_series(ch, data_ws, info["ONs_MD"], info["ONs_TIR"], "TIR observada",
                          markers_only=True, color="1F2A44")
            if "TIR_Curva" in d.columns and d["TIR_Curva"].notna().any():
                _add_xy_series(ch, data_ws, info["ONs_MD"], info["ONs_TIRCurva"], "Curva NS",
                              line_only=True, color="C8963E")
            ws.add_chart(ch, f"F{chart_anchor_row}")
    except Exception:
        pass
    chart_anchor_row += 19

    # 2) Curva Hard Dollar: un scatter por legislación (GD/AL/BOPREAL)
    try:
        df_hd_c = ctx.get("df_hd")
        if df_hd_c is not None and not df_hd_c.empty and {"Grupo", "MD", "TIR"}.issubset(df_hd_c.columns):
            ch2 = _scatter_chart("Soberanos Hard Dollar — TIR vs MD por legislación",
                                "Modified Duration (años)", "TIR (%)")
            colores_hd = {"GD": "2563EB", "AL": "DC2626", "BOPREAL": "059669"}
            col_cursor = 4
            for grupo, dg in df_hd_c.groupby("Grupo"):
                dg = dg.dropna(subset=["MD", "TIR"]).sort_values("MD")
                if dg.empty:
                    continue
                info_g = _write_columns(data_ws, [f"HD_{grupo}_MD", f"HD_{grupo}_TIR"],
                                        [dg["MD"].tolist(), dg["TIR"].tolist()],
                                        start_row=1, start_col=col_cursor)
                _add_xy_series(ch2, data_ws, info_g[f"HD_{grupo}_MD"], info_g[f"HD_{grupo}_TIR"], grupo,
                              markers_only=True, color=colores_hd.get(grupo, "64748B"))
                col_cursor += 2
            ws.add_chart(ch2, f"F{chart_anchor_row}")
    except Exception:
        pass
    chart_anchor_row += 19

    # 3) Breakeven CER + referencia REM
    try:
        df_cer_c = ctx.get("df_carry_cer")
        if df_cer_c is not None and not df_cer_c.empty and {"MD", "Breakeven_%anual"}.issubset(df_cer_c.columns):
            d = df_cer_c.dropna(subset=["MD", "Breakeven_%anual"]).sort_values("MD")
            cols = ["BE_CER_MD", "BE_CER_Val"]
            data = [d["MD"].tolist(), d["Breakeven_%anual"].tolist()]
            rem_raw = ctx.get("rem_val_raw")
            if rem_raw is not None:
                cols += ["BE_CER_RefMD", "BE_CER_Ref"]
                data += [[d["MD"].min(), d["MD"].max()], [rem_raw, rem_raw]]
            info3 = _write_columns(data_ws, cols, data, start_row=1, start_col=20)
            ch3 = _scatter_chart("Breakeven inflación — CER vs Fija", "Modified Duration (años)", "Breakeven (% anual)")
            _add_xy_series(ch3, data_ws, info3["BE_CER_MD"], info3["BE_CER_Val"], "Breakeven CER",
                          markers_only=True, color="7C3AED")
            if rem_raw is not None:
                _add_xy_series(ch3, data_ws, info3["BE_CER_RefMD"], info3["BE_CER_Ref"], "REM 12m",
                              line_only=True, color="C8963E")
            ws.add_chart(ch3, f"F{chart_anchor_row}")
    except Exception:
        pass
    chart_anchor_row += 19

    # 4) Breakeven DL + referencia de devaluación
    try:
        df_dl_c = ctx.get("df_carry_dl")
        if df_dl_c is not None and not df_dl_c.empty and {"MD", "Breakeven_%anual"}.issubset(df_dl_c.columns):
            d = df_dl_c.dropna(subset=["MD", "Breakeven_%anual"]).sort_values("MD")
            cols = ["BE_DL_MD", "BE_DL_Val"]
            data = [d["MD"].tolist(), d["Breakeven_%anual"].tolist()]
            dev_raw = ctx.get("rofex_12m_raw")
            if dev_raw is not None:
                cols += ["BE_DL_RefMD", "BE_DL_Ref"]
                data += [[d["MD"].min(), d["MD"].max()], [dev_raw, dev_raw]]
            info4 = _write_columns(data_ws, cols, data, start_row=1, start_col=26)
            ch4 = _scatter_chart("Breakeven devaluación — DL vs Fija", "Modified Duration (años)", "Breakeven (% anual)")
            _add_xy_series(ch4, data_ws, info4["BE_DL_MD"], info4["BE_DL_Val"], "Breakeven DL",
                          markers_only=True, color="EA580C")
            if dev_raw is not None:
                _add_xy_series(ch4, data_ws, info4["BE_DL_RefMD"], info4["BE_DL_Ref"], "Deval. 12m",
                              line_only=True, color="C8963E")
            ws.add_chart(ch4, f"F{chart_anchor_row}")
    except Exception:
        pass
    chart_anchor_row += 19

    # 5) Spread histórico por legislación (serie de tiempo)
    try:
        df_sp_hist = ctx.get("df_spread_hist")
        if df_sp_hist is not None and not df_sp_hist.empty and "Date" in df_sp_hist.columns:
            year_cols = [c for c in df_sp_hist.columns if c != "Date"]
            cols = ["Spread_Date"] + [f"Spread_{c}" for c in year_cols]
            fechas = [pd.Timestamp(x).to_pydatetime() for x in df_sp_hist["Date"]]
            data = [fechas] + [df_sp_hist[c].tolist() for c in year_cols]
            info5 = _write_columns(data_ws, cols, data, start_row=1, start_col=32)
            lc = LineChart()
            lc.title = "Spread por legislación (AL/AE − GD) — últimas ruedas"
            lc.style = 2
            lc.x_axis.title = "Fecha"
            lc.y_axis.title = "Spread (pp de TIR)"
            lc.x_axis.axPos = "b"
            lc.y_axis.axPos = "l"
            lc.x_axis.delete = False
            lc.y_axis.delete = False
            lc.x_axis.number_format = "dd/mm/yy"
            lc.height = 8.5
            lc.width = 16.0
            date_col, dr0, dr1 = info5["Spread_Date"]
            xref = Reference(data_ws, min_col=date_col, min_row=dr0, max_row=dr1)
            for yc in year_cols:
                col_i, r0, r1 = info5[f"Spread_{yc}"]
                yref = Reference(data_ws, min_col=col_i, min_row=r0 - 1, max_row=r1)
                lc.add_data(yref, titles_from_data=True)
            lc.set_categories(xref)
            ws.add_chart(lc, f"F{chart_anchor_row}")
    except Exception:
        pass
    chart_anchor_row += 20

    # ---------- Tablas de sensibilidad (heatmap con formato condicional) ----------
    try:
        chart_anchor_row = _add_sensitivity_table(ws, ctx.get("df_carry_cer"), "CER", chart_anchor_row, start_col=6)
    except Exception:
        pass
    try:
        chart_anchor_row = _add_sensitivity_table(ws, ctx.get("df_carry_dl"), "DL", chart_anchor_row, start_col=6)
    except Exception:
        pass

    # ---------- Hoja 2: ONs ----------
    ws2 = wb.create_sheet("ONs - Valor Relativo")
    _xls_write_table(ws2, ctx.get("df_ons"), 1, "Obligaciones Negociables — TIR, MD y valor relativo vs curva NS")

    # ---------- Hoja 3: Bonos Hard Dollar ----------
    ws3 = wb.create_sheet("Hard Dollar")
    nr = _xls_write_table(ws3, ctx.get("df_hd"), 1, "Soberanos Hard Dollar — snapshot y cheap/rich por legislación")
    _xls_write_table(ws3, ctx.get("df_spread"), nr, "Spread por legislación (AL/AE − GD), z-score 252 ruedas")

    # ---------- Hoja 4: Breakeven & Carry CER ----------
    ws4 = wb.create_sheet("Breakeven CER")
    _xls_write_table(ws4, ctx.get("df_carry_cer"), 1,
                     f"Inflación — CER vs Fija (escenario: {ctx.get('infl_esc','—')}% anual)")

    # ---------- Hoja 5: Breakeven & Carry DL ----------
    ws5 = wb.create_sheet("Breakeven DL")
    _xls_write_table(ws5, ctx.get("df_carry_dl"), 1,
                     f"Devaluación — DL vs Fija (escenario: {ctx.get('deval_esc','—')}% anual)")

    # ---------- Hoja 6: Auditoría de exclusiones ----------
    ws6 = wb.create_sheet("Auditoría - Excluidos")
    _xls_write_table(ws6, ctx.get("df_excluidos"), 1,
                     "Instrumentos excluidos del fit de curva y breakevens (MD<0,3 o TIR fuera de rango)")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


st.set_page_config(page_title="Renta Fija PRO — Curvas, Spreads, Breakevens y Carry Trade", layout="wide")

st.markdown(
    f"""
    <div style="padding:14px 20px;border-radius:12px;
         background:linear-gradient(90deg,{COLORS['primary']},#31456e);color:white;">
      <span style="font-size:1.45rem;font-weight:700;">📊 Renta Fija Argentina — Panel PRO</span><br>
      <span style="opacity:.85;">Curvas ONs y soberanos · valor relativo · spreads por legislación · breakevens y carry trade vs REM y ROFEX</span>
    </div>
    """,
    unsafe_allow_html=True,
)
st.caption("Herramienta de análisis. No constituye recomendación de inversión.")

with st.sidebar:
    st.header("⚙️ Configuración")
    api_key = st.text_input("Alphacast API Key", value="", type="password")
    dataset_id = st.number_input(
        "Dataset principal (ONs, Bonos HD y Soberanos en pesos)",
        value=int(DEFAULT_DATASET_ID), step=1,
        help="Un único dataset cubre todo el universo: 'market segment' distingue "
             "Corporate/Sovereign y 'coupon structure' distingue CER/Fija/DL/Dual.",
    )

    st.divider()
    st.subheader("Filtros de ONs")
    lamina_max = st.number_input("Lámina mínima (≤)", value=float(DEFAULT_LAMINA_MAX), step=1.0)
    md_max = st.number_input("MD máxima (≤)", value=float(DEFAULT_MD_MAX), step=0.1)
    rating = st.text_input("Calificación exacta (vacío = todas)", value="AAA(arg)")

    st.divider()
    st.subheader("Breakevens y Carry Trade")
    ds_rem = st.number_input("Dataset REM (BCRA)", value=int(DATASET_REM), step=1)
    ds_fut = st.number_input("Dataset Futuros ROFEX", value=int(DATASET_FUTUROS), step=1)

if not api_key.strip():
    st.warning("Ingresá tu Alphacast API Key para descargar los datasets.")
    st.stop()

with st.spinner("Descargando dataset principal..."):
    raw = download_dataset(api_key.strip(), int(dataset_id))

df_norm = normalize_dataset(raw)
latest_df, most_recent_date = latest_snapshot(df_norm)

tab_ons, tab_bonos, tab_be, tab_insights, tab_export = st.tabs(
    ["🏢 ONs", "💵 Bonos Hard Dollar", "⚖️ Breakevens & Carry Trade", "🎯 Insights", "📥 Exportar Informe"]
)

# ---------------------------------------------------------------------
# TAB 1 — ONs
# ---------------------------------------------------------------------
with tab_ons:
    st.markdown("Editá el universo de ONs (tickers, calificación, lámina). La salida se filtra por lámina, MD y calificación.")

    if "ons_table" not in st.session_state:
        st.session_state["ons_table"] = default_ons_df()

    left, right = st.columns([2, 1])
    with right:
        st.subheader("Carga rápida")
        up = st.file_uploader("Importar ONs desde CSV", type=["csv"])
        if up is not None:
            try:
                du = pd.read_csv(up)
                needed = {"Ticker", "Calificacion", "Lamina_Minima"}
                if not needed.issubset(du.columns):
                    st.error("El CSV debe tener columnas: Ticker, Calificacion, Lamina_Minima")
                else:
                    du["Lamina_Minima"] = pd.to_numeric(du["Lamina_Minima"], errors="coerce")
                    st.session_state["ons_table"] = du[list(needed)].copy()
                    st.success("ONs importadas.")
            except Exception as e:
                st.error(f"No se pudo leer el CSV: {e}")
        if st.button("Restaurar ONs precargadas"):
            st.session_state["ons_table"] = default_ons_df()

    with left:
        st.subheader("Universo de ONs")
        edited = st.data_editor(
            st.session_state["ons_table"], num_rows="dynamic", use_container_width=True, hide_index=True,
            column_config={
                "Ticker": st.column_config.TextColumn("Ticker", required=True),
                "Calificacion": st.column_config.TextColumn("Calificación"),
                "Lamina_Minima": st.column_config.NumberColumn("Lámina mínima", min_value=0.0, step=1.0),
            },
            key="ons_editor",
        )
        st.session_state["ons_table"] = edited.copy()

    df_ons = build_ons_table(latest_df, st.session_state["ons_table"])
    df_ons["Lamina_Minima"] = pd.to_numeric(df_ons.get("Lamina_Minima"), errors="coerce")
    df_ons_f = df_ons[
        (pd.to_numeric(df_ons["Lamina_Minima"], errors="coerce") <= float(lamina_max))
        & (pd.to_numeric(df_ons["MD"], errors="coerce") <= float(md_max))
    ].copy()
    if rating.strip():
        df_ons_f = df_ons_f[df_ons_f["Calificacion"].astype(str) == rating.strip()]

    fig_ons, params_ons = (None, {})
    if not df_ons_f.empty:
        fig_ons, params_ons = plot_curve_pro(df_ons_f, "Curva de ONs — TIR vs MD (fit Nelson-Siegel, banda ±50 bps)",
                                             fecha=most_recent_date)
        df_ons_f = cheap_rich(df_ons_f, params_ons.get("global"))

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Instrumentos", len(df_ons_f))
        c2.metric("TIR promedio", f"{df_ons_f['TIR'].mean():.2f}%")
        c3.metric("MD promedio", f"{df_ons_f['MD'].mean():.2f}")
        best = df_ons_f.loc[df_ons_f["Residuo_bps"].idxmax()] if df_ons_f["Residuo_bps"].notna().any() else None
        c4.metric("Más barato vs curva", best["Ticker"] if best is not None else "—",
                  f"+{best['Residuo_bps']:.0f} bps" if best is not None else "")

        if fig_ons is not None:
            st.plotly_chart(fig_ons, use_container_width=True)

        st.subheader("Valor relativo (cheap / rich vs curva NS)")
        # Flag de liquidez: marca los de volumen bajo (percentil 25 del universo filtrado)
        if "Volumen" in df_ons_f.columns and df_ons_f["Volumen"].notna().any():
            umbral_liq = df_ons_f["Volumen"].quantile(0.25)
            df_ons_f["Liquidez"] = np.where(df_ons_f["Volumen"] <= umbral_liq, "🔸 Baja", "✅ OK")
        cols_show = [c for c in ["Ticker", "TIR", "MD", "Convexidad", "Paridad", "Calificacion",
                                 "Lamina_Minima", "Volumen", "Liquidez", "TIR_Curva", "Residuo_bps",
                                 "Señal"] if c in df_ons_f.columns]
        st.dataframe(df_ons_f[cols_show].sort_values("Residuo_bps", ascending=False),
                     use_container_width=True, height=380)
        st.caption("Residuo = TIR observada − TIR teórica de la curva. Positivo (>50 bps): rinde por encima de "
                   "sus comparables de igual duration → candidato *barato*. Negativo: *caro*. "
                   "**Columna Volumen** (última rueda) y flag de liquidez: un residuo alto en un bono 🔸 de baja "
                   "liquidez suele ser espejismo — el precio no refleja mercado real. Cruzá siempre residuo con volumen.")
    else:
        st.info("No hay ONs que pasen los filtros actuales.")

# ---------------------------------------------------------------------
# TAB 2 — Bonos HD
# ---------------------------------------------------------------------
with tab_bonos:
    all_bonds = BONOS_GD + BONOS_AL + BOPREAL
    sel = st.multiselect("Seleccionar bonos", options=all_bonds, default=all_bonds)

    lb = latest_df[latest_df["Ticker"].isin(sel)].copy()
    lb["Grupo"] = lb["Ticker"].apply(classify_bono)

    # Aviso de tickers seleccionados que no están en el snapshot (antes desaparecían sin explicación)
    faltantes = [t for t in sel if t not in set(lb["Ticker"])]
    if faltantes:
        st.info(f"Sin dato en el último snapshot ({pd.Timestamp(most_recent_date).strftime('%d/%m/%Y') if most_recent_date is not None else '—'}): "
                f"**{', '.join(faltantes)}**. Puede ser por falta de cotización ese día, ticker no listado en el dataset, "
                f"o especie no operada. No se grafican para no inducir a error.")

    if not lb.empty:
        # Fit propio por grupo y señal cheap/rich (igual criterio que ONs)
        fig_b, params_b = plot_curve_pro(
            lb, "Soberanos Hard Dollar — TIR vs MD por legislación", fecha=most_recent_date,
            group_col="Grupo", group_colors={"GD": COLORS["gd"], "AL": COLORS["al"], "BOPREAL": COLORS["bopreal"]},
            fit_per_group=True,
        )
        if fig_b is not None:
            st.plotly_chart(fig_b, use_container_width=True)

        # Señal cheap/rich: para GD y AL usamos su propio fit de grupo; BOPREAL aparte
        partes = []
        for g, dfg in lb.groupby("Grupo"):
            p = (params_b or {}).get(g)
            partes.append(cheap_rich(dfg, p) if p is not None else dfg.assign(TIR_Curva=np.nan, Residuo_bps=np.nan, Señal="⚪ s/fit"))
        lb_cr = pd.concat(partes, ignore_index=True)
        lb_cr["TIR/MD"] = (lb_cr["TIR"] / lb_cr["MD"]).round(2)

        st.subheader("Snapshot con valor relativo (cheap/rich vs curva de su grupo)")
        cols_hd = [c for c in ["Ticker", "Grupo", "TIR", "MD", "Convexidad", "Paridad", "Volumen",
                               "TIR/MD", "TIR_Curva", "Residuo_bps", "Señal"] if c in lb_cr.columns]
        st.dataframe(lb_cr[cols_hd].sort_values(["Grupo", "Residuo_bps"], ascending=[True, False]),
                     use_container_width=True, height=340)
        st.caption("Señal cheap/rich = TIR observada − TIR teórica de la curva **de su propia legislación** "
                   "(GD contra GD, AL contra AL). Un AL barato no es comparable con un GD barato: son curvas distintas. "
                   "**BOPREAL es deuda del BCRA, no del Tesoro** — su riesgo de crédito es del Banco Central "
                   "(reservas), no soberano puro; no se lo mezcla en la curva GD/AL.")
    else:
        st.info("Ningún bono seleccionado tiene datos en el último snapshot.")

    st.subheader("Spread por legislación: AL/AE − GD")
    spreads = spread_series(df_norm)
    stats = spread_stats(spreads)
    if not stats.empty:
        st.dataframe(stats, use_container_width=True, height=240)
        st.caption("Z-score sobre 252 ruedas. Un spread muy por encima de su media histórica (z ≥ 1.5) indica que la ley "
                   "local está inusualmente castigada: si se espera compresión, el bono AL captura ese exceso de rendimiento. "
                   "Un z ≤ −1.5 sugiere que el premio por ley NY está barato en términos relativos. "
                   "Los pares se arman por año de vencimiento; si falta una pata (p.ej. no hay AL46 contra GD46), ese año no aparece.")
        fig_sp = plot_spread_history(spreads)
        if fig_sp is not None:
            st.plotly_chart(fig_sp, use_container_width=True)
    else:
        st.info("No hay pares AL/GD suficientes para calcular spreads históricos.")

# ---------------------------------------------------------------------
# TAB 3 — Breakevens & Carry Trade
# ---------------------------------------------------------------------
with tab_be:
    st.markdown("### ⚖️ Curva en pesos: ¿cobertura o tasa fija?")
    st.caption(
        "Tres preguntas, en orden: (1) **Breakeven** — ¿qué inflación/devaluación tenés que "
        "creer para estar indiferente entre cubrirte (CER/DL) y quedarte en tasa fija? "
        "(2) **Escenario** — dado tu supuesto (REM/ROFEX por defecto, o el tuyo), ¿qué pata "
        "rinde más? (3) **Instrumento** — a igual duration, ¿cuál captura mejor el trade?"
    )

    # ---- Universo depurado (filtros de calidad + sin defaulteados) ----
    latest_sob = classify_soberanos_pesos(latest_df, max_md=float(md_max) if md_max else 6.0)

    if latest_sob.empty or not {"Ticker", "TIR", "MD", "Clase"}.issubset(latest_sob.columns):
        st.info("No se encontraron soberanos en pesos operables en el dataset (revisar filtros/segmento).")
        st.stop()

    fija_all = latest_sob[latest_sob["Clase"] == "Fija"].dropna(subset=["MD", "TIR"])
    cer_all = latest_sob[latest_sob["Clase"] == "CER"].dropna(subset=["MD", "TIR"])
    dl_all = latest_sob[latest_sob["Clase"] == "DL"].dropna(subset=["MD", "TIR"])
    fija_df, fija_susp = split_curve_outliers(fija_all)
    cer_df, cer_susp = split_curve_outliers(cer_all)
    dl_df, dl_susp = split_curve_outliers(dl_all)
    susp_total = pd.concat([cer_susp.assign(Clase="CER"), fija_susp.assign(Clase="Fija"),
                            dl_susp.assign(Clase="DL")], ignore_index=True) if any(
        [not cer_susp.empty, not fija_susp.empty, not dl_susp.empty]) else pd.DataFrame()
    params_fija, _ = fit_ns(fija_df["MD"], fija_df["TIR"]) if len(fija_df) >= 5 else (None, None)

    if not susp_total.empty:
        with st.expander(f"⚠️ {len(susp_total)} instrumento(s) excluidos del fit por MD<0.3 o TIR fuera de rango "
                         f"(precio/cupón a revisar) — clic para ver", expanded=False):
            cols_s = [c for c in ["Ticker", "Clase", "MD", "TIR", "Paridad"] if c in susp_total.columns]
            st.dataframe(susp_total[cols_s].sort_values("MD"), use_container_width=True, height=180)
            st.caption("Estos NO entran a la curva Nelson-Siegel ni a los breakevens porque distorsionan el tramo "
                       "corto (la TIR anualizada de un bono a días de vencer es numéricamente inestable). "
                       "Se muestran para que los audites, no se ocultan.")

    counts = latest_sob["Clase"].value_counts()
    with st.expander(f"Universo operable: Fija {counts.get('Fija',0)} · CER {counts.get('CER',0)} · "
                     f"DL {counts.get('DL',0)} · Dual {counts.get('Dual',0)} — clic para reclasificar", expanded=False):
        st.caption("Se excluyeron especies en USD (C/D) y bonos de reestructuración 2005/2010 "
                   "(PARP/DICP/CUAP: son performing, pero ilíquidos y de estructura atípica, no defaulteados). "
                   "Reclasificá acá si querés forzar un caso borde (p.ej. mover un Dual a CER según cómo esperás "
                   "que termine indexando).")
        for k in ["CER", "Fija", "DL", "Dual"]:
            auto = sorted(latest_sob.loc[latest_sob["Clase"] == k, "Ticker"].astype(str).unique())
            pick = st.multiselect(f"{k}", options=sorted(latest_sob["Ticker"].astype(str).unique()),
                                  default=auto, key=f"pick_{k}")
            latest_sob.loc[latest_sob["Ticker"].astype(str).isin(pick), "Clase"] = k

    # ---- Descarga de expectativas externas (REM / ROFEX) ----
    rem_val, rem_fecha = None, None
    try:
        with st.spinner("REM (BCRA)…"):
            raw_rem = download_dataset(api_key.strip(), int(ds_rem))
        rem_val, rem_fecha = get_rem_inflacion_12m(raw_rem)
    except Exception as e:
        st.warning(f"No se pudo procesar el REM (dataset {ds_rem}): {e}")

    fut_curve, fut_spot, fut_fecha, fut_diag = pd.DataFrame(), None, None, {"ok": False, "motivo": "no descargado"}
    fut_error_tecnico = None
    try:
        with st.spinner("Futuros ROFEX…"):
            raw_fut = download_dataset(api_key.strip(), int(ds_fut))
        fut_curve, fut_spot, fut_fecha, fut_diag = get_futuros_curva(raw_fut)
    except Exception as e:
        fut_error_tecnico = f"{type(e).__name__}: {e}"

    # Devaluación anualizada de referencia a ~12m.
    # 1º intento: curva de futuros ROFEX. Si no está disponible, se usa la
    # devaluación implícita de los propios bonos DL (breakeven fija-vs-DL a 1 año).
    rofex_12m = interp_futuros_annual(fut_curve, 365) if (fut_curve is not None and not fut_curve.empty) else None
    rofex_12m = None if (rofex_12m is None or pd.isna(rofex_12m)) else round(float(rofex_12m), 1)
    if rofex_12m is not None:
        deval_source = "ROFEX (curva de futuros)"
    else:
        rofex_12m = deval_implicita_desde_bonos(fija_df, dl_df, params_fija, horizon_days=365)
        deval_source = "Devaluación implícita de bonos DL (fallback)"

    # ---- Panel de ESCENARIOS (editable) ----
    st.markdown("#### 🎚️ Escenario macro (12 meses)")
    e1, e2, e3 = st.columns([1, 1, 2])
    with e1:
        infl_default = float(rem_val) if rem_val is not None else 25.0
        infl_esc = st.number_input("Inflación esperada (% i.a.)", value=round(infl_default, 1), step=0.5,
                                   help=f"Default = REM {('%.1f%%' % rem_val) if rem_val is not None else 'N/D'} "
                                        f"({pd.Timestamp(rem_fecha).strftime('%m/%Y') if rem_fecha is not None else '—'}). "
                                        "Editá para probar tu propio escenario (bull/base/bear).")
    with e2:
        deval_default = float(rofex_12m) if rofex_12m is not None else 25.0
        deval_esc = st.number_input("Devaluación esperada (% i.a.)", value=round(deval_default, 1), step=0.5,
                                    help=f"Default = ROFEX 12m {('%.1f%%' % rofex_12m) if rofex_12m is not None else 'N/D'}. "
                                         "Es la devaluación anualizada implícita en la curva de futuros.")
    with e3:
        st.caption(f"**Referencias de mercado hoy** — REM inflación 12m: "
                   f"**{('%.1f%%' % rem_val) if rem_val is not None else 'N/D'}**  ·  "
                   f"Devaluación 12m: **{('%.1f%%' % rofex_12m) if rofex_12m is not None else 'N/D'}**  ·  "
                   f"Spot: **{('%.1f' % fut_spot) if fut_spot else '—'}**")
        st.caption(f"Fuente devaluación: _{deval_source}_.")
        if fut_error_tecnico is not None:
            st.caption(f"⚙️ Falla técnica al bajar futuros (dataset {ds_fut}): `{fut_error_tecnico}`. "
                       f"No es un juicio de calidad del mercado, es un problema de pipeline; "
                       f"se usó el fallback de bonos DL.")
        elif not fut_diag.get("ok", False):
            st.caption(f"ℹ️ Curva ROFEX sin tenors utilizables: {fut_diag.get('motivo', 'no disponible')}.")

    # ---- Motor de carry por escenario ----
    carry_cer = carry_scenario(cer_df, fija_df, "CER", infl_esc, params_fija) if not cer_df.empty else pd.DataFrame()
    carry_dl = carry_scenario(dl_df, fija_df, "DL", deval_esc, params_fija) if not dl_df.empty else pd.DataFrame()

    # =====================  TABLERO DE DECISIÓN  =====================
    st.markdown("#### 📌 Tablero de decisión")
    k1, k2, k3, k4 = st.columns(4)

    # NOTA: el delta se define como (breakeven - referencia). Signo negativo = el
    # mercado pricea MENOS inflación/deval que la referencia → la cobertura luce
    # barata (bueno para quien quiere cubrirse). Usamos delta_color="inverse" para
    # que ese caso (delta<0) se muestre en verde de forma coherente con la lectura.
    be_cer_short = carry_cer[carry_cer["MD"] <= 1.5]["Breakeven_%anual"].mean() if not carry_cer.empty else np.nan
    if pd.notna(be_cer_short) and rem_val is not None:
        delta_cer = be_cer_short - rem_val  # breakeven vs REM
        k1.metric("Breakeven inflación (tramo corto)", f"{be_cer_short:.1f}%",
                  f"{delta_cer:+.1f} pp vs REM ({rem_val:.1f}%)", delta_color="inverse")
    else:
        k1.metric("Breakeven inflación (corto)", f"{be_cer_short:.1f}%" if pd.notna(be_cer_short) else "—")

    if not carry_cer.empty and "Excess_pp" in carry_cer.columns:
        best_cer = carry_cer.loc[carry_cer["Excess_pp"].idxmax()]
        k2.metric("Mejor CER (bajo tu escenario)", best_cer["Ticker"],
                  f"{best_cer['Excess_pp']:+.1f} pp vs Fija")
    else:
        k2.metric("Mejor CER", "—")

    be_dl_short = carry_dl[carry_dl["MD"] <= 1.5]["Breakeven_%anual"].mean() if not carry_dl.empty else np.nan
    if pd.notna(be_dl_short) and rofex_12m is not None:
        delta_dl = be_dl_short - rofex_12m  # breakeven vs referencia de devaluación
        ref_txt = "ROFEX" if fut_diag.get("ok") else "impl."
        k3.metric("Breakeven devaluación (corto)", f"{be_dl_short:.1f}%",
                  f"{delta_dl:+.1f} pp vs {ref_txt} ({rofex_12m:.1f}%)", delta_color="inverse")
    else:
        k3.metric("Breakeven devaluación (corto)", f"{be_dl_short:.1f}%" if pd.notna(be_dl_short) else "—")

    # Señal dominante CER vs Fija bajo escenario
    if not carry_cer.empty and "Excess_pp" in carry_cer.columns:
        med_excess = carry_cer["Excess_pp"].median()
        senal = "Cobertura CER" if med_excess > 1 else ("Tasa Fija" if med_excess < -1 else "Indiferente")
        k4.metric("Señal dominante (CER vs Fija)", senal, f"mediana {med_excess:+.1f} pp")
    else:
        k4.metric("Señal dominante", "—")

    # Insight en prosa (autogenerado)
    if pd.notna(be_cer_short) and rem_val is not None:
        gap = rem_val - be_cer_short
        if gap > 2:
            st.success(f"**Lectura:** el mercado pricea {be_cer_short:.1f}% de inflación en el tramo corto, "
                       f"por **debajo** del REM ({rem_val:.1f}%). Si la inflación converge al REM, el **CER está barato** "
                       f"y supera a la tasa fija: el breakeven es fácil de batir.")
        elif gap < -2:
            st.success(f"**Lectura:** el breakeven corto ({be_cer_short:.1f}%) está **por encima** del REM "
                       f"({rem_val:.1f}%). El mercado ya paga cara la cobertura: si el REM acierta, la **tasa fija "
                       f"rinde más** en ese tramo.")
        else:
            st.info(f"**Lectura:** breakeven corto ({be_cer_short:.1f}%) ≈ REM ({rem_val:.1f}%). No hay carry claro "
                    f"por inflación en el tramo corto; la decisión pasa por duration, liquidez y tu visión propia.")

    st.divider()

    # =====================  (A) BREAKEVEN + (B) CARRY  =====================
    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("##### 🔥 Inflación — CER vs Fija")
        if not carry_cer.empty:
            fig = plot_breakeven(
                carry_cer.rename(columns={"Breakeven_%anual": "Breakeven_%anual"}),
                "Inflación breakeven implícita por plazo", COLORS["cer"],
                ref_value=rem_val, ref_label=f"REM 12m: {rem_val:.1f}%" if rem_val is not None else None,
            )
            # superponer escenario del usuario si difiere del REM
            if fig is not None and abs(infl_esc - (rem_val or infl_esc)) > 0.05:
                fig.add_hline(y=infl_esc, line_dash="dot", line_color=COLORS["cer"],
                              annotation_text=f"Tu escenario: {infl_esc:.1f}%",
                              annotation_font_color=COLORS["cer"])
            if fig is not None:
                st.plotly_chart(fig, use_container_width=True)

            st.markdown(f"**Carry bajo escenario = {infl_esc:.1f}% de inflación**")
            cols_show = [c for c in ["Ticker", "MD", "TIR_Real", "TIR_Fija_interp", "Breakeven_%anual",
                                     "Ret_cobertura_%", "Ret_fija_%", "Excess_pp", "Margen_vs_BE_pp",
                                     "Recomendación"] if c in carry_cer.columns]
            st.dataframe(carry_cer[cols_show].sort_values("Excess_pp", ascending=False)
                         if "Excess_pp" in carry_cer.columns else carry_cer[cols_show],
                         use_container_width=True, height=300)
            st.caption("**Breakeven_%anual**: inflación de indiferencia CER↔Fija. "
                       "**Excess_pp**: retorno esperado de la cobertura menos el de la fija bajo tu escenario "
                       "(positivo = conviene CER). **Margen_vs_BE_pp**: cuánta inflación de más asumís sobre el breakeven.")
        else:
            st.info("Faltan bonos CER y/o Fija operables para calcular breakevens.")

    with col_b:
        st.markdown("##### 💱 Devaluación — DL vs Fija")
        if not carry_dl.empty:
            fig = plot_breakeven(
                carry_dl, "Devaluación breakeven implícita por plazo", COLORS["dl"],
                ref_value=rofex_12m,
                ref_label=(f"Deval. 12m ({'ROFEX' if fut_diag.get('ok') else 'implícita'}): {rofex_12m:.1f}%"
                           if rofex_12m is not None else None),
            )
            if fig is not None and rofex_12m is not None and abs(deval_esc - rofex_12m) > 0.05:
                fig.add_hline(y=deval_esc, line_dash="dot", line_color=COLORS["dl"],
                              annotation_text=f"Tu escenario: {deval_esc:.1f}%",
                              annotation_font_color=COLORS["dl"])
            if fig is not None:
                st.plotly_chart(fig, use_container_width=True)

            st.markdown(f"**Carry bajo escenario = {deval_esc:.1f}% de devaluación**")
            cols_show = [c for c in ["Ticker", "MD", "TIR_DL", "TIR_Fija_interp", "Breakeven_%anual",
                                     "Ret_cobertura_%", "Ret_fija_%", "Excess_pp", "Margen_vs_BE_pp",
                                     "Recomendación"] if c in carry_dl.columns]
            st.dataframe(carry_dl[cols_show].sort_values("Excess_pp", ascending=False)
                         if "Excess_pp" in carry_dl.columns else carry_dl[cols_show],
                         use_container_width=True, height=300)
            st.caption("Análogo al CER pero en dólar-linked: TIR_DL es tasa en USD sobre la trayectoria del "
                       "A3500. El retorno esperado en pesos es (1+TIR_DL)·(1+devaluación)−1, comparado contra la "
                       "tasa fija de igual duration. La referencia de devaluación surge de "
                       f"_{deval_source}_.")
        else:
            st.info("Faltan bonos DL y/o Fija operables para calcular breakevens.")

    st.divider()

    # =====================  (C) SENSIBILIDAD: MAPAS DE CALOR  =====================
    st.markdown("##### 🌡️ Sensibilidad — excess return por letra y escenario")
    st.caption(
        "Cada celda es el **excess return** (en pp) de tener la cobertura vs. la tasa fija comparable, "
        "para una inflación/devaluación anual dada. **Verde**: conviene la cobertura (CER/DL). **Rojo**: "
        "conviene la tasa fija. La **✕** en cada fila es el breakeven de ese bono (donde el excess = 0); "
        "la línea vertical punteada es tu escenario. Leé una fila de izquierda a derecha para ver a partir "
        "de qué nivel de inflación/devaluación conviene cubrirse con esa letra."
    )

    hcol1, hcol2 = st.columns(2)
    with hcol1:
        if not carry_cer.empty:
            rng = st.slider("Rango de inflación anual (%) — eje del heatmap CER",
                            min_value=0.0, max_value=80.0,
                            value=(max(0.0, (rem_val or 25) - 15), (rem_val or 25) + 20),
                            step=1.0, key="rng_cer")
            fig_h = build_heatmap(carry_cer, "CER", rng[0], rng[1], n=30, escenario_marker=infl_esc)
            if fig_h is not None:
                st.plotly_chart(fig_h, use_container_width=True)
        else:
            st.info("Sin datos CER para el heatmap.")

    with hcol2:
        if not carry_dl.empty:
            base_dev = rofex_12m if rofex_12m is not None else 25.0
            rng2 = st.slider("Rango de devaluación anual (%) — eje del heatmap DL",
                             min_value=-10.0, max_value=120.0,
                             value=(max(-10.0, base_dev - 20), base_dev + 30),
                             step=1.0, key="rng_dl")
            fig_h2 = build_heatmap(carry_dl, "DL", rng2[0], rng2[1], n=30, escenario_marker=deval_esc)
            if fig_h2 is not None:
                st.plotly_chart(fig_h2, use_container_width=True)
        else:
            st.info("Sin datos DL para el heatmap.")

    st.divider()

    # =====================  (D) VALOR RELATIVO: curvas superpuestas  =====================
    st.markdown("##### 📈 Curvas en pesos superpuestas (valor relativo a igual duration)")
    d3 = latest_sob[latest_sob["Clase"].isin(["CER", "Fija", "DL", "Dual"])].copy()
    if not d3.empty:
        fig3, _ = plot_curve_pro(d3, "TIR vs MD por clase (fit NS por grupo)", fecha=most_recent_date,
                                 group_col="Clase",
                                 group_colors={"CER": COLORS["cer"], "Fija": COLORS["fija"],
                                               "DL": COLORS["dl"], "Dual": COLORS["dual"]},
                                 fit_per_group=True)
        if fig3 is not None:
            st.plotly_chart(fig3, use_container_width=True)
        st.caption("La curva CER está en tasa **real** y la Fija en tasa **nominal**: la brecha vertical entre "
                   "ambas ES la inflación breakeven a cada duration. La DL está en tasa **USD**; su brecha vs Fija "
                   "es la devaluación breakeven.")

    with st.expander("Ver curva de futuros ROFEX / REM crudos (auditoría de datos)"):
        cc, cd = st.columns(2)
        with cc:
            st.markdown("**Curva de devaluación implícita (por plazo)**")
            if fut_curve is not None and not fut_curve.empty:
                st.dataframe(fut_curve, use_container_width=True, height=240)
                st.caption(f"Spot {fut_spot:.1f} · tenors válidos {fut_diag.get('tenors_validos','?')}/"
                           f"{fut_diag.get('tenors_totales','?')} · fuente: {deval_source}")
            else:
                st.caption(f"Sin curva de futuros utilizable: {fut_diag.get('motivo','—')}. "
                           f"Se usó devaluación implícita de bonos DL como referencia.")
        with cd:
            st.markdown("**REM — inflación esperada (últimos)**")
            if rem_val is not None:
                st.caption(f"IPC nivel general · 'Próx. 12 meses' · mediana **{rem_val:.1f}%** "
                           f"({pd.Timestamp(rem_fecha).strftime('%m/%Y')})")

    st.caption("⚠️ Supuestos: mismo horizonte de tenencia, sin reinversión de cupón ni costos de transacción; "
               "TIR anual efectiva. El breakeven es un punto de indiferencia, no una predicción; la expectativa "
               "externa (REM/ROFEX) puede errar. Esto es análisis, no recomendación de inversión.")

# ---------------------------------------------------------------------
# TAB 4 — Insights automáticos
# ---------------------------------------------------------------------
with tab_insights:
    st.subheader("🎯 Resumen ejecutivo para el asesor")
    st.caption("Síntesis automática de las 3 pestañas. Reacciona a los filtros y escenarios que elijas.")
    bullets = []
    caveats = []

    # --- Hard Dollar: forma de curva GD ---
    lb_all = latest_df[latest_df["Ticker"].isin(BONOS_GD)].dropna(subset=["MD", "TIR"])
    if len(lb_all) >= 3:
        corto = lb_all.nsmallest(2, "MD")["TIR"].mean()
        largo = lb_all.nlargest(2, "MD")["TIR"].mean()
        pend = largo - corto
        forma = "invertida (más tasa en el tramo corto → estrés de corto plazo)" if pend < -0.5 \
            else ("empinada (premio por extender duration)" if pend > 0.5 else "plana")
        bullets.append(f"**Hard Dollar — curva GD {forma}.** Tramo corto ≈ {corto:.1f}%, largo ≈ {largo:.1f}% "
                       f"(pendiente {pend:+.1f} pp).")

    # --- Spread legislación ---
    try:
        stats_i = spread_stats(spread_series(df_norm))
        if not stats_i.empty:
            ext = stats_i.loc[stats_i["Z_Score"].abs().idxmax()]
            if abs(ext["Z_Score"]) >= 1.5:
                bullets.append(f"**Spread ley 20{int(ext['Numero'])} en extremo histórico** "
                               f"(z = {ext['Z_Score']:+.2f}; {ext['Spread_Actual']:.2f} pp vs prom. "
                               f"{ext['Spread_Prom_252']:.2f}). {ext['Señal']}.")
            else:
                bullets.append("**Spread AL/GD** dentro de rangos normales (|z| < 1.5 en todos los pares): "
                               "sin señal fuerte de arbitraje por legislación hoy.")
    except Exception:
        pass

    # --- ONs cheap/rich con chequeo de liquidez ---
    try:
        if "df_ons_f" in dir() and isinstance(df_ons_f, pd.DataFrame) and "Residuo_bps" in df_ons_f.columns \
                and df_ons_f["Residuo_bps"].notna().any():
            top = df_ons_f.nlargest(3, "Residuo_bps")
            def _tag(r):
                liq = getattr(r, "Liquidez", "")
                return f"{r.Ticker} (+{r.Residuo_bps:.0f} bps{', 🔸baja liq.' if 'Baja' in str(liq) else ''})"
            names = ", ".join(_tag(r) for r in top.itertuples())
            bullets.append(f"**ONs más baratas vs curva:** {names}. "
                           f"Los marcados 🔸 tienen volumen bajo: el exceso de TIR puede ser espejismo de precio.")
    except Exception:
        pass

    # --- Carry CER: usar EXACTAMENTE el breakeven del tramo corto del Tablero (consistencia) ---
    try:
        if "carry_cer" in dir() and isinstance(carry_cer, pd.DataFrame) and not carry_cer.empty:
            be_short = carry_cer[carry_cer["MD"] <= 1.5]["Breakeven_%anual"].mean()
            be_full = carry_cer["Breakeven_%anual"].mean()
            ref_rem = f" vs REM {rem_val:.1f}%" if ("rem_val" in dir() and rem_val is not None) else ""
            lado = ""
            if "rem_val" in dir() and rem_val is not None and pd.notna(be_short):
                lado = (" → CER barato, conviene cubrirse" if be_short < rem_val - 1
                        else (" → CER caro, conviene tasa fija" if be_short > rem_val + 1 else " → neutral"))
            bullets.append(f"**Carry CER vs Fija:** breakeven de inflación tramo corto (MD≤1,5) "
                           f"**{be_short:.1f}%**{ref_rem}{lado}. Curva completa: {be_full:.1f}% promedio. "
                           f"(El Tablero usa el de tramo corto; este promedio full-curve es solo referencia agregada.)")
    except Exception:
        pass

    # --- CAVEATS de datos sucios (lo que el diagnóstico pidió que Insights no ocultara) ---
    try:
        if "susp_total" in dir() and isinstance(susp_total, pd.DataFrame) and not susp_total.empty:
            tk = ", ".join(susp_total.sort_values("MD")["Ticker"].head(4))
            caveats.append(f"**{len(susp_total)} instrumento(s) en pesos excluidos del análisis** por MD<0,3 o TIR "
                           f"fuera de rango ({tk}…): rendimientos inestables cerca del vencimiento. No operar sobre esos números.")
    except Exception:
        pass
    try:
        if "fut_error_tecnico" in dir() and fut_error_tecnico is not None:
            caveats.append(f"**Curva de futuros ROFEX no disponible** (falla técnica de pipeline): la devaluación "
                           f"de referencia se estimó desde bonos DL. Verificá antes de operar tipo de cambio.")
        elif "fut_diag" in dir() and not fut_diag.get("ok", True):
            caveats.append("**Curva ROFEX sin tenors utilizables**: devaluación de referencia estimada desde bonos DL.")
    except Exception:
        pass
    # ONs con TIR negativa (dato a auditar)
    try:
        if "df_ons_f" in dir() and isinstance(df_ons_f, pd.DataFrame) and (df_ons_f["TIR"] < 0).any():
            negs = ", ".join(df_ons_f[df_ons_f["TIR"] < 0]["Ticker"].astype(str).head(4))
            caveats.append(f"**ONs con TIR negativa** ({negs}): precio/cupón a auditar; arrastran el tramo corto de la curva.")
    except Exception:
        pass

    if bullets:
        for b in bullets:
            st.markdown(f"- {b}")
    else:
        st.info("Cargá datos en las otras pestañas para generar insights.")

    if caveats:
        st.markdown("##### ⚠️ Caveats de datos (revisar antes de operar)")
        for c in caveats:
            st.markdown(f"- {c}")

    st.divider()
    st.markdown(
        """
        **Cómo leer este panel (guía rápida):**
        - *Cheap/rich*: un bono ±50 bps fuera de la curva NS no es orden de compra/venta; suele reflejar liquidez,
          lámina mínima o riesgo idiosincrático. Es un **filtro**, no una señal. Cruzá siempre con volumen.
        - *Spread AL−GD*: mide el premio de la ley local. Su z-score dice si está caro/barato **en términos relativos**,
          no si el país mejora o empeora.
        - *Breakeven*: inflación/devaluación de **indiferencia** entre cubrirse y tasa fija. El "tramo corto" (MD≤1,5)
          es lo que muestra el Tablero; el "promedio full-curve" que a veces se cita es una agregación distinta —no los mezcles.
        - *Carry por escenario*: el excess return depende de TU supuesto macro. Si no tenés visión, el default es el
          consenso (REM/ROFEX), y ahí el carry ≈ 0 por construcción. El valor está en jugar escenarios propios.
        """
    )

# ---------------------------------------------------------------------
# TAB 5 — Exportar Informe (Excel)
# ---------------------------------------------------------------------
with tab_export:
    st.markdown("### 📥 Informe de Renta Fija en Excel")
    st.caption(
        "Descarga un .xlsx con el resumen ejecutivo, el tablero de decisión, los insights y caveats, "
        "y las tablas completas de ONs, Hard Dollar, Breakeven/Carry CER y DL. Es una **foto** del "
        "estado actual de la app (mismo escenario y filtros que tengas seleccionados ahora), no un "
        "modelo con fórmulas — para volver a calcular con otro escenario, generá un nuevo informe."
    )

    # Arma el contexto tomando lo que cada pestaña ya calculó en esta misma corrida del script.
    # Uso defensivo (dir()/try-except) porque alguna pestaña puede no haber corrido si no hay datos.
    ctx = {}
    ctx["fecha_datos"] = pd.Timestamp(most_recent_date).strftime("%d/%m/%Y") if most_recent_date is not None else "—"

    try:
        ctx["infl_esc"] = infl_esc
        ctx["infl_es_rem"] = (rem_val is not None and abs(infl_esc - rem_val) < 0.05)
        ctx["rem_val_txt"] = f"{rem_val:.1f}%" if rem_val is not None else "N/D"
        ctx["rem_val_raw"] = float(rem_val) if rem_val is not None else None
        ctx["deval_esc"] = deval_esc
        ctx["deval_source"] = deval_source
        ctx["rofex_12m_raw"] = float(rofex_12m) if rofex_12m is not None else None
    except Exception:
        pass

    try:
        ctx["be_cer_short"] = round(float(be_cer_short), 2) if pd.notna(be_cer_short) else None
        ctx["delta_cer_txt"] = (f"{delta_cer:+.1f} pp vs REM ({rem_val:.1f}%)"
                               if "delta_cer" in dir() and rem_val is not None else None)
    except Exception:
        pass
    try:
        ctx["best_cer_txt"] = f"{best_cer['Ticker']} ({best_cer['Excess_pp']:+.1f} pp vs Fija)"
    except Exception:
        pass
    try:
        ctx["be_dl_short"] = round(float(be_dl_short), 2) if pd.notna(be_dl_short) else None
        ctx["delta_dl_txt"] = (f"{delta_dl:+.1f} pp vs {ref_txt} ({rofex_12m:.1f}%)"
                               if "delta_dl" in dir() and rofex_12m is not None else None)
    except Exception:
        pass
    try:
        ctx["senal_dominante"] = f"{senal} (mediana {med_excess:+.1f} pp)"
    except Exception:
        pass
    try:
        if pd.notna(be_cer_short) and rem_val is not None:
            gap_txt = rem_val - be_cer_short
            if gap_txt > 2:
                ctx["lectura"] = (f"El mercado pricea {be_cer_short:.1f}% de inflación en el tramo corto, por "
                                 f"debajo del REM ({rem_val:.1f}%): el CER luce barato, conviene cobertura si "
                                 f"el REM se cumple.")
            elif gap_txt < -2:
                ctx["lectura"] = (f"El breakeven corto ({be_cer_short:.1f}%) está por encima del REM "
                                 f"({rem_val:.1f}%): la cobertura ya está cara, la tasa fija rinde más si el "
                                 f"REM se cumple.")
            else:
                ctx["lectura"] = (f"Breakeven corto ({be_cer_short:.1f}%) ≈ REM ({rem_val:.1f}%): sin carry "
                                 f"claro por inflación; la decisión pasa por duration y liquidez.")
    except Exception:
        pass

    ctx["bullets"] = bullets if "bullets" in dir() else []
    ctx["caveats"] = caveats if "caveats" in dir() else []

    try:
        ctx["df_ons"] = df_ons_f if "df_ons_f" in dir() and isinstance(df_ons_f, pd.DataFrame) else pd.DataFrame()
    except Exception:
        ctx["df_ons"] = pd.DataFrame()
    try:
        ctx["df_hd"] = lb_cr if "lb_cr" in dir() and isinstance(lb_cr, pd.DataFrame) else pd.DataFrame()
    except Exception:
        ctx["df_hd"] = pd.DataFrame()
    try:
        ctx["df_spread"] = stats if "stats" in dir() and isinstance(stats, pd.DataFrame) else pd.DataFrame()
    except Exception:
        ctx["df_spread"] = pd.DataFrame()
    try:
        ctx["df_spread_hist"] = (_pivot_spread_history(spreads)
                                if "spreads" in dir() and isinstance(spreads, pd.DataFrame) else pd.DataFrame())
    except Exception:
        ctx["df_spread_hist"] = pd.DataFrame()
    try:
        ctx["df_carry_cer"] = carry_cer if "carry_cer" in dir() and isinstance(carry_cer, pd.DataFrame) else pd.DataFrame()
    except Exception:
        ctx["df_carry_cer"] = pd.DataFrame()
    try:
        ctx["df_carry_dl"] = carry_dl if "carry_dl" in dir() and isinstance(carry_dl, pd.DataFrame) else pd.DataFrame()
    except Exception:
        ctx["df_carry_dl"] = pd.DataFrame()
    try:
        ctx["df_excluidos"] = susp_total if "susp_total" in dir() and isinstance(susp_total, pd.DataFrame) else pd.DataFrame()
    except Exception:
        ctx["df_excluidos"] = pd.DataFrame()

    # ---- Preview de lo que va a llevar el informe ----
    p1, p2, p3 = st.columns(3)
    p1.metric("ONs en el informe", len(ctx.get("df_ons", pd.DataFrame())))
    p2.metric("Soberanos HD", len(ctx.get("df_hd", pd.DataFrame())))
    p3.metric("CER + DL con carry", len(ctx.get("df_carry_cer", pd.DataFrame())) + len(ctx.get("df_carry_dl", pd.DataFrame())))

    try:
        excel_bytes = build_excel_report(ctx)
        nombre_archivo = f"renta_fija_informe_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.download_button(
            "⬇️ Descargar informe Excel (.xlsx)",
            data=excel_bytes,
            file_name=nombre_archivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption(
            "Hojas incluidas: **Resumen** (tablero + insights + caveats), **ONs - Valor Relativo**, "
            "**Hard Dollar** (snapshot + spread por legislación), **Breakeven CER**, **Breakeven DL**, "
            "y **Auditoría - Excluidos** (instrumentos fuera del fit, para que quede documentado por qué)."
        )
    except Exception as e:
        st.error(f"No se pudo generar el informe: {e}")
